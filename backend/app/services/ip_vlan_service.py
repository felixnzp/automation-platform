from __future__ import annotations

import ipaddress
import json
import re
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from threading import Lock
import time

from netmiko import ConnectHandler
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.device import Device
from app.models.ip_vlan_task import IpVlanChangeLog

MAC_REGEX = re.compile(r"([0-9a-fA-F]{4}[-:.][0-9a-fA-F]{4}[-:.][0-9a-fA-F]{4}|[0-9a-fA-F]{2}(?::[0-9a-fA-F]{2}){5})")
INTERFACE_REGEX = re.compile(r"((?:GigabitEthernet|XGigabitEthernet|GE|XGE|Eth-Trunk)\S+)", re.IGNORECASE)
INTERFACE_BLOCK_RE = re.compile(r"^interface\s+(\S+)\s*$", re.IGNORECASE)
_TOPOLOGY_CACHE: dict | None = None
_LOOKUP_CACHE: dict[str, tuple[float, object]] = {}
_LOOKUP_CACHE_LOCK = Lock()


class LocateTraceError(ValueError):
    def __init__(self, message: str, logs: list[dict] | None = None):
        super().__init__(message)
        self.logs = logs or []


def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _cache_get(key: str):
    now = time.time()
    with _LOOKUP_CACHE_LOCK:
        item = _LOOKUP_CACHE.get(key)
        if not item:
            return None
        expires_at, value = item
        if expires_at <= now:
            _LOOKUP_CACHE.pop(key, None)
            return None
        return value


def _cache_set(key: str, value, ttl_seconds: int):
    with _LOOKUP_CACHE_LOCK:
        _LOOKUP_CACHE[key] = (time.time() + max(1, ttl_seconds), value)


def _safe_int(value, default=22) -> int:
    try:
        return int(value)
    except Exception:
        return default


def _normalize_mac(mac_text: str) -> str:
    text = str(mac_text or "").strip().lower().replace(":", "").replace("-", "").replace(".", "")
    if len(text) != 12:
        return str(mac_text or "").strip().lower()
    return f"{text[0:4]}-{text[4:8]}-{text[8:12]}"


def _validate_ip(target_ip: str) -> str:
    try:
        return str(ipaddress.ip_address(str(target_ip or "").strip()))
    except Exception as exc:
        raise ValueError("IP不合法") from exc


def _infer_role(device: Device) -> str:
    text = f"{device.name or ''} {device.group_name or ''} {device.location or ''}".lower()
    if any(key in text for key in ["router", "璺敱"]):
        return "router"
    if any(key in text for key in ["core", "鏍稿績", "csw"]):
        return "core_switch"
    return "access_switch"


def _connect_device(device: Device, timeout: int = 20):
    return ConnectHandler(
        device_type=device.device_type or "huawei",
        host=device.ip,
        username=device.username,
        password=device.password,
        port=_safe_int(device.port, 22),
        fast_cli=False,
        use_keys=False,
        allow_agent=False,
        conn_timeout=max(timeout, 5),
        banner_timeout=max(timeout, 5),
        auth_timeout=max(timeout, 5),
    )


def _normalize_interface_for_config(interface_name: str) -> str:
    """
    Normalize interface name for Huawei VRP config commands.
    Examples:
      GE0/0/3  -> GigabitEthernet0/0/3
      XGE0/0/13 -> XGigabitEthernet0/0/13
      Eth-Trunk2 -> Eth-Trunk2
    """
    text = str(interface_name or "").strip()
    if not text:
        return ""

    # Remove accidental spaces: "GigabitEthernet 0/0/3" -> "GigabitEthernet0/0/3"
    text = re.sub(r"^(GigabitEthernet|XGigabitEthernet)\\s+(\\d)", r"\\1\\2", text, flags=re.IGNORECASE)
    text = text.replace(" ", "")

    upper = text.upper()
    if upper.startswith("ETH-TRUNK"):
        return text
    if upper.startswith("GIGABITETHERNET") or upper.startswith("XGIGABITETHERNET"):
        return text
    if upper.startswith("XGE"):
        return f"XGigabitEthernet{text[3:]}"
    if upper.startswith("GE"):
        return f"GigabitEthernet{text[2:]}"
    return text


def _huawei_save(conn, timeout: int = 90) -> str:
    """
    Persist running config to startup config on Huawei VRP devices.
    Usually requires interactive confirmation: `save` -> `y`.
    """
    try:
        try:
            conn.exit_config_mode()
        except Exception:
            # Best-effort: return to user view (similar to Ctrl+Z).
            try:
                conn.send_command_timing("return", read_timeout=10)
            except Exception:
                pass

        output = conn.send_command_timing("save", read_timeout=timeout)
        if re.search(r"\\[Y/N\\]|\\(y/n\\)|are you sure|continue\\?", output, re.IGNORECASE):
            output2 = conn.send_command_timing("y", read_timeout=timeout)
            output = f"{output}\\n{output2}"
        return output
    except Exception as exc:
        raise ValueError(f"保存配置失败: {exc}") from exc


def _choose_access_uplink_port(topology_row: dict | None, fallback_port: str = "") -> str:
    if not topology_row:
        return str(fallback_port or "").strip()
    eth_trunk = str(topology_row.get("eth_trunk") or "").strip()
    neighbor_intf = str(topology_row.get("neighbor_intf") or "").strip()
    if eth_trunk and eth_trunk.upper() != "NONE":
        return eth_trunk
    if neighbor_intf and neighbor_intf.upper() != "NONE":
        return neighbor_intf
    return str(fallback_port or "").strip()


def _select_core_switch(db: Session) -> Device | None:
    rows = db.query(Device).order_by(Device.id.asc()).all()
    exact_name = next((row for row in rows if str(row.name or "").strip().lower() == "sz-csw"), None)
    if exact_name:
        return exact_name

    exact_ip = next((row for row in rows if str(row.ip or "").strip() == "10.18.100.1"), None)
    if exact_ip:
        return exact_ip

    preferred = []
    fallback = []
    for row in rows:
        role = _infer_role(row)
        text = f"{row.name or ''} {row.group_name or ''} {row.location or ''}".lower()
        if role == "core_switch":
            if any(key in text for key in ["csw", "core", "鏍稿績"]) and not any(key in text for key in ["router", "璺敱"]):
                preferred.append(row)
            else:
                fallback.append(row)
    if preferred:
        return preferred[0]
    if fallback:
        return fallback[0]
    return rows[0] if rows else None


def _infer_floor_text(device: Device) -> str:
    for text in [str(device.name or ""), str(device.group_name or ""), str(device.location or "")]:
        match = re.search(r"(17F|18F)", text, re.IGNORECASE)
        if match:
            return match.group(1).upper()
    return ""


def _infer_floor_from_target_ip(target_ip: str) -> str:
    text = str(target_ip or "").strip()
    if text.startswith("10.17."):
        return "17F"
    if text.startswith("10.18."):
        return "18F"
    return ""


def _list_access_switches(db: Session, target_ip: str = "") -> list[Device]:
    rows = db.query(Device).order_by(Device.id.asc()).all()
    access_rows = [row for row in rows if _infer_role(row) == "access_switch"]
    floor = _infer_floor_from_target_ip(target_ip)
    if not floor:
        return access_rows

    floor_matched = [row for row in access_rows if _infer_floor_text(row) == floor]
    return floor_matched or access_rows


def _run_show(conn, command: str, timeout: int = 20) -> str:
    return conn.send_command(command, read_timeout=max(timeout, 10)) or ""


def _parse_arp_mac(output: str) -> str | None:
    match = MAC_REGEX.search(output or "")
    if not match:
        return None
    return _normalize_mac(match.group(1))


def _parse_arp_interface(output: str) -> str | None:
    for line in (output or "").splitlines():
        if "eth-trunk" not in line.lower() and "gigabitethernet" not in line.lower() and "ge" not in line.lower():
            continue
        match = INTERFACE_REGEX.search(line)
        if match:
            return match.group(1)
    return None


def _extract_arp_section(full_output: str, target_ip: str) -> str:
    lines = []
    for line in (full_output or "").splitlines():
        if str(target_ip) in line:
            lines.append(line)
    return "\n".join(lines)


def _parse_port_from_mac(output: str) -> str | None:
    for line in (output or "").splitlines():
        lower = line.lower()
        if "vlanif" in lower or "cpu" in lower:
            continue
        match = INTERFACE_REGEX.search(line)
        if match:
            return match.group(1)
    return None


def _normalize_name_key(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(text or "").lower())


def _normalize_interface_key(text: str) -> str:
    return re.sub(r"[^a-z0-9/]", "", str(text or "").lower())


def _topology_file_path() -> Path:
    return Path(__file__).resolve().parents[3] / "devices" / "NetDevices.xlsx"


def _load_topology_cache() -> dict:
    global _TOPOLOGY_CACHE
    if _TOPOLOGY_CACHE is not None:
        return _TOPOLOGY_CACHE

    by_eth_trunk: dict[str, dict] = {}
    by_neighbor_intf: dict[str, dict] = {}
    by_name: dict[str, dict] = {}
    by_ip: dict[str, dict] = {}
    path = _topology_file_path()
    if not path.exists():
        _TOPOLOGY_CACHE = {
            "by_eth_trunk": by_eth_trunk,
            "by_neighbor_intf": by_neighbor_intf,
            "by_name": by_name,
            "by_ip": by_ip,
        }
        return _TOPOLOGY_CACHE

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            index_map = {str(value or "").strip(): idx for idx, value in enumerate(header)}
            if not {"NAME", "IP", "Eth-Trunk", "Neighbor_Intf"}.issubset(index_map):
                continue
            for row in rows:
                name = str(row[index_map["NAME"]] or "").strip()
                ip = str(row[index_map["IP"]] or "").strip()
                eth_trunk = str(row[index_map["Eth-Trunk"]] or "").strip()
                neighbor_intf = str(row[index_map["Neighbor_Intf"]] or "").strip()
                if not name and not ip:
                    continue
                item = {
                    "name": name,
                    "ip": ip,
                    "eth_trunk": eth_trunk,
                    "neighbor_intf": neighbor_intf,
                }
                if name:
                    by_name[_normalize_name_key(name)] = item
                if ip:
                    by_ip[ip] = item
                if eth_trunk and eth_trunk.upper() != "NONE":
                    by_eth_trunk[_normalize_interface_key(eth_trunk)] = item
                if neighbor_intf and neighbor_intf.upper() != "NONE":
                    by_neighbor_intf[_normalize_interface_key(neighbor_intf)] = item
    finally:
        wb.close()

    _TOPOLOGY_CACHE = {
        "by_eth_trunk": by_eth_trunk,
        "by_neighbor_intf": by_neighbor_intf,
        "by_name": by_name,
        "by_ip": by_ip,
    }
    return _TOPOLOGY_CACHE


def _find_topology_row(core_uplink: str, core_uplink_desc: str = "") -> dict | None:
    cache = _load_topology_cache()
    interface_key = _normalize_interface_key(core_uplink)
    if interface_key:
        row = cache["by_eth_trunk"].get(interface_key)
        if row:
            return row
        row = cache["by_neighbor_intf"].get(interface_key)
        if row:
            return row
    desc_key = _normalize_name_key(core_uplink_desc)
    if desc_key:
        row = cache["by_name"].get(desc_key)
        if row:
            return row
    return None


def _find_access_switch_by_row(db: Session, topology_row: dict | None) -> Device | None:
    if not topology_row:
        return None
    ip = str(topology_row.get("ip") or "").strip()
    if ip:
        by_ip = db.query(Device).filter(Device.ip == ip).first()
        if by_ip:
            return by_ip
    name_key = _normalize_name_key(topology_row.get("name") or "")
    if name_key:
        for row in _list_access_switches(db):
            if _normalize_name_key(row.name or "") == name_key:
                return row
    return None


def _find_access_switch_by_hint(db: Session, hint: str) -> Device | None:
    if not hint:
        return None

    normalized_hint = _normalize_name_key(hint)
    rows = _list_access_switches(db)
    for row in rows:
        if _normalize_name_key(row.name) == normalized_hint:
            return row
    for row in rows:
        if normalized_hint and normalized_hint in _normalize_name_key(row.name):
            return row
    return None


def _ordered_access_switches(db: Session, target_ip: str, hinted_switch: Device | None) -> list[Device]:
    switches = _list_access_switches(db, target_ip)
    if not hinted_switch:
        return switches
    return [hinted_switch]


def _format_interface_for_display(interface_name: str) -> str:
    text = str(interface_name or "").strip()
    patterns = [
        (r"^(GigabitEthernet)(\d+/\d+/\d+)$", r"\1 \2"),
        (r"^(XGigabitEthernet)(\d+/\d+/\d+)$", r"\1 \2"),
        (r"^(GE)(\d+/\d+/\d+)$", r"GigabitEthernet \2"),
        (r"^(XGE)(\d+/\d+/\d+)$", r"XGigabitEthernet \2"),
        (r"^(Eth-Trunk)(\d+)$", r"\1 \2"),
    ]
    for pattern, replacement in patterns:
        if re.match(pattern, text, re.IGNORECASE):
            return re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    return text


def _compact_interface_name(interface_name: str) -> str:
    text = str(interface_name or "").strip()
    patterns = [
        (r"^(GigabitEthernet)\s+(\d+/\d+/\d+)$", r"\1\2"),
        (r"^(XGigabitEthernet)\s+(\d+/\d+/\d+)$", r"\1\2"),
        (r"^(Eth-Trunk)\s+(\d+)$", r"\1\2"),
    ]
    for pattern, replacement in patterns:
        if re.match(pattern, text, re.IGNORECASE):
            return re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    return text.replace(" ", "")


def _interface_command_variants(interface_name: str) -> list[str]:
    candidates = []
    for item in [
        str(interface_name or "").strip(),
        _format_interface_for_display(interface_name),
        _compact_interface_name(_format_interface_for_display(interface_name)),
    ]:
        if item and item not in candidates:
            candidates.append(item)
    return candidates


def _looks_like_cli_error(output: str) -> bool:
    text = str(output or "").lower()
    return any(marker in text for marker in ["error:", "unrecognized", "incomplete command", "wrong parameter", "too many parameters"])


def _run_interface_show(conn, prefix: str, interface_name: str, timeout: int, logs: list[dict], device_name: str) -> tuple[str, str]:
    last_output = ""
    last_command = ""
    for candidate in _interface_command_variants(interface_name):
        command = f"{prefix} {candidate}"
        output = _run_show(conn, command, timeout)
        logs.append({"device": device_name, "command": command, "output": output[-2000:]})
        last_command = command
        last_output = output
        if output and not _looks_like_cli_error(output):
            return command, output
    return last_command, last_output


def _extract_description_name(output: str) -> str:
    match = re.search(r"^\s*description\s+(.+)$", output or "", re.IGNORECASE | re.MULTILINE)
    if not match:
        return ""
    return str(match.group(1) or "").strip()


def _build_mac_lookup_commands(mac: str) -> list[str]:
    normalized = _normalize_mac(mac)
    compact = normalized.replace("-", "")
    dotted = f"{compact[0:4]}.{compact[4:8]}.{compact[8:12]}" if len(compact) == 12 else normalized
    return [
        f"display mac-address | include {normalized}",
        f"display mac-address {normalized}",
        f"display mac-address {dotted}",
    ]


def _normalize_mac_keyword(keyword: str) -> str:
    return re.sub(r"[^0-9a-fA-F]", "", str(keyword or "")).lower()


def _extract_mac_matches(output: str) -> list[dict]:
    matches: list[dict] = []
    seen: set[tuple[str, str]] = set()
    for line in (output or "").splitlines():
        mac_match = MAC_REGEX.search(line or "")
        iface_match = INTERFACE_REGEX.search(line or "")
        if not mac_match:
            continue
        mac = _normalize_mac(mac_match.group(1))
        interface_name = iface_match.group(1) if iface_match else ""
        key = (mac, interface_name)
        if key in seen:
            continue
        seen.add(key)
        matches.append({"mac": mac, "interface_name": interface_name, "line": line.strip()})
    return matches


def _build_exact_uplink_checks(core_uplink: str, core_output: str, access_output: str, new_vlan: int | None = None) -> list[dict]:
    rows = []
    if core_uplink:
        core_allowed_vlans = _extract_allowed_vlans_text(core_output)
        access_allowed_vlans = _extract_allowed_vlans_text(access_output)
        core_has_vlan = _interface_vlan_allowed(core_output, new_vlan) if new_vlan else False
        access_has_vlan = _interface_vlan_allowed(access_output, new_vlan) if new_vlan else False
        rows.append(
            {
                "interface_name": f"鏍稿績:{core_uplink}",
                "port_type": "trunk",
                "description": "鏍稿績浜ゆ崲鏈轰笂鑱斿彛",
                "allowed_vlans": core_allowed_vlans or "-",
                "has_vlan": core_has_vlan,
                "already_allows": core_has_vlan,
                "requires_add": False if not new_vlan else not core_has_vlan,
                "planned_commands": [] if not new_vlan or core_has_vlan else _build_uplink_add_command(core_uplink, "trunk", new_vlan),
            }
        )
        rows.append(
            {
                "interface_name": f"鎺ュ叆:{core_uplink}",
                "port_type": "trunk",
                "description": "鎺ュ叆浜ゆ崲鏈轰笂鑱斿彛",
                "allowed_vlans": access_allowed_vlans or "-",
                "has_vlan": access_has_vlan,
                "already_allows": access_has_vlan,
                "requires_add": False if not new_vlan else not access_has_vlan,
                "planned_commands": [] if not new_vlan or access_has_vlan else _build_uplink_add_command(core_uplink, "trunk", new_vlan),
            }
        )
    return rows


def _build_exact_uplink_checks_dual(core_uplink: str, access_uplink: str, core_output: str, access_output: str, new_vlan: int | None = None) -> list[dict]:
    rows = []
    core_uplink = str(core_uplink or "").strip()
    access_uplink = str(access_uplink or "").strip() or core_uplink
    if core_uplink:
        core_allowed_vlans = _extract_allowed_vlans_text(core_output)
        access_allowed_vlans = _extract_allowed_vlans_text(access_output)
        core_has_vlan = _interface_vlan_allowed(core_output, new_vlan) if new_vlan else False
        access_has_vlan = _interface_vlan_allowed(access_output, new_vlan) if new_vlan else False
        rows.append(
            {
                "interface_name": f"核心:{core_uplink}",
                "device_side": "core",
                "device_interface": core_uplink,
                "port_type": "trunk",
                "description": "核心交换机互联口",
                "allowed_vlans": core_allowed_vlans or "-",
                "has_vlan": core_has_vlan,
                "already_allows": core_has_vlan,
                "requires_add": False if not new_vlan else not core_has_vlan,
                "planned_commands": [] if not new_vlan or core_has_vlan else _build_uplink_add_command(core_uplink, "trunk", new_vlan),
            }
        )
        rows.append(
            {
                "interface_name": f"接入:{access_uplink}",
                "device_side": "access",
                "device_interface": access_uplink,
                "port_type": "trunk",
                "description": "接入交换机上联口",
                "allowed_vlans": access_allowed_vlans or "-",
                "has_vlan": access_has_vlan,
                "already_allows": access_has_vlan,
                "requires_add": False if not new_vlan else not access_has_vlan,
                "planned_commands": [] if not new_vlan or access_has_vlan else _build_uplink_add_command(access_uplink, "trunk", new_vlan),
            }
        )
    return rows


def _fetch_uplink_snapshot(db: Session, locate_data: dict, logs: list[dict] | None = None) -> dict:
    core_uplink = str(locate_data.get("core_uplink_interface") or "").strip()
    access_uplink = str(locate_data.get("access_uplink_interface") or "").strip() or core_uplink
    access_switch_data = locate_data.get("access_switch") or {}
    core_switch_data = locate_data.get("core_switch") or {}
    access_switch = db.query(Device).filter(Device.id == int(access_switch_data.get("id") or 0)).first()
    core_switch = db.query(Device).filter(Device.id == int(core_switch_data.get("id") or 0)).first()

    if not core_uplink or not access_switch or not core_switch:
        return {
            "core_uplink_interface": core_uplink,
            "access_uplink_interface": access_uplink,
            "core_allowed_vlans": "",
            "access_allowed_vlans": "",
            "uplink_plan": [],
        }

    worker_logs: list[dict] = [] if logs is None else logs
    cache_key = f"uplink-snapshot:{core_switch.id}:{access_switch.id}:{_compact_interface_name(core_uplink)}:{_compact_interface_name(access_uplink)}"
    cached = _cache_get(cache_key)
    if cached is not None:
        if logs is not None:
            logs.append({"device": "system", "command": "uplink-snapshot-cache", "output": f"鍛戒腑閾捐矾缂撳瓨 {cache_key}"})
        return dict(cached)

    def fetch_core():
        return _fetch_interface_config(core_switch, core_uplink, 20, 120, worker_logs)

    def fetch_access():
        return _fetch_interface_config(access_switch, access_uplink, 20, 120, worker_logs)

    with ThreadPoolExecutor(max_workers=2) as executor:
        core_future = executor.submit(fetch_core)
        access_future = executor.submit(fetch_access)
        core_uplink_output = core_future.result()
        access_uplink_output = access_future.result()

    uplink_plan = _build_exact_uplink_checks_dual(core_uplink, access_uplink, core_uplink_output, access_uplink_output, None)
    result = {
        "core_uplink_interface": core_uplink,
        "access_uplink_interface": access_uplink,
        "core_allowed_vlans": uplink_plan[0]["allowed_vlans"] if len(uplink_plan) > 0 else "",
        "access_allowed_vlans": uplink_plan[1]["allowed_vlans"] if len(uplink_plan) > 1 else "",
        "uplink_plan": uplink_plan,
    }
    _cache_set(cache_key, result, 120)
    return result


def _parse_port_profile(output: str) -> dict:
    text = output or ""
    lower = text.lower()

    port_type = "access"
    if "port link-type trunk" in lower:
        port_type = "trunk"
    elif "port link-type hybrid" in lower:
        port_type = "hybrid"

    current_vlan = "1"
    match_default = re.search(r"port\s+default\s+vlan\s+(\d+)", lower)
    if match_default:
        current_vlan = match_default.group(1)

    description = ""
    match_desc = re.search(r"^\s*description\s+(.+)$", text, re.IGNORECASE | re.MULTILINE)
    if match_desc:
        description = match_desc.group(1).strip()

    return {
        "port_type": port_type,
        "current_vlan": current_vlan,
        "description": description,
    }


def _parse_port_status(brief_output: str) -> str:
    text = (brief_output or "").lower()
    if "up" in text:
        return "up"
    if "down" in text:
        return "down"
    return "unknown"


def _is_forbidden_port(interface_name: str, profile: dict) -> tuple[bool, str]:
    iface = str(interface_name or "").lower()
    port_type = str(profile.get("port_type", "")).lower()
    description = str(profile.get("description", "")).lower()

    if "eth-trunk" in iface:
        return True, "褰撳墠绔彛鏄仛鍚堝彛锛屼笉鍏佽鐩存帴淇敼"
    if port_type in {"trunk", "hybrid"}:
        return True, "当前端口不是 access 口"
    if any(key in description for key in ["uplink", "涓婅仈", "core", "鏍稿績", "trunk", "姹囪仛"]):
        return True, "当前端口疑似上联口，不允许直接修改"
    return False, ""


def _parse_interface_blocks(config_text: str) -> dict[str, list[str]]:
    blocks: dict[str, list[str]] = {}
    current_name = ""

    for raw_line in (config_text or "").splitlines():
        line = raw_line.rstrip("\r")
        match = INTERFACE_BLOCK_RE.match(line.strip())
        if match:
            current_name = match.group(1)
            blocks[current_name] = [line]
            continue
        if current_name:
            blocks[current_name].append(line)
    return blocks


def _interface_vlan_allowed(block_text: str, new_vlan: int) -> bool:
    lower = block_text.lower()
    if " port trunk allow-pass vlan all" in lower or " port hybrid tagged vlan all" in lower:
        return True

    patterns = [
        r"port\s+trunk\s+allow-pass\s+vlan\s+(.+)",
        r"port\s+hybrid\s+tagged\s+vlan\s+(.+)",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, lower):
            if _vlan_in_spec(match.group(1), new_vlan):
                return True
    return False


def _extract_allowed_vlans_text(block_text: str) -> str:
    lower = str(block_text or "").lower()
    if "port trunk allow-pass vlan all" in lower:
        return "all"

    match = re.search(r"port\s+trunk\s+allow-pass\s+vlan\s+(.+)", str(block_text or ""), re.IGNORECASE)
    if match:
        return " ".join(str(match.group(1) or "").split())
    return ""


def _fetch_interface_config(device: Device, interface_name: str, timeout: int, cache_ttl: int, logs: list[dict] | None = None) -> str:
    cache_key = f"iface-config:{device.id}:{_compact_interface_name(interface_name)}"
    cached = _cache_get(cache_key)
    if cached is not None:
        if logs is not None:
            logs.append(
                {
                    "device": device.name,
                    "command": f"display current-configuration interface {interface_name}",
                    "output": f"[cache-hit]\n{str(cached)[-1800:]}",
                }
            )
        return str(cached)

    conn = _connect_device(device, timeout=timeout)
    try:
        _cmd, output = _run_interface_show(
            conn,
            "display current-configuration interface",
            interface_name,
            timeout,
            logs if logs is not None else [],
            device.name,
        )
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass

    _cache_set(cache_key, output, cache_ttl)
    return output


def _validate_uplink_pass_through(db: Session, locate_data: dict, new_vlan: int, logs: list[dict] | None = None) -> dict:
    core_uplink = str(locate_data.get("core_uplink_interface") or "").strip()
    access_uplink = str(locate_data.get("access_uplink_interface") or "").strip() or core_uplink
    access_switch_data = locate_data.get("access_switch") or {}
    core_switch_data = locate_data.get("core_switch") or {}
    access_switch = db.query(Device).filter(Device.id == int(access_switch_data.get("id") or 0)).first()
    core_switch = db.query(Device).filter(Device.id == int(core_switch_data.get("id") or 0)).first()
    cache_key = f"uplink-validate:{int(core_switch_data.get('id') or 0)}:{int(access_switch_data.get('id') or 0)}:{_compact_interface_name(core_uplink)}:{_compact_interface_name(access_uplink)}:{new_vlan}"

    cached = _cache_get(cache_key)
    if cached is not None:
        if logs is not None:
            logs.append({"device": "system", "command": "uplink-validate-cache", "output": f"鍛戒腑鏀鹃€氭牎楠岀紦瀛?{cache_key}"})
        return dict(cached)

    if not core_uplink or not access_switch or not core_switch:
        return {
            "core_uplink_interface": core_uplink,
            "access_uplink_interface": access_uplink,
            "core_allowed_vlans": "",
            "access_allowed_vlans": "",
            "core_has_vlan": False,
            "access_has_vlan": False,
            "can_modify": False,
            "reason": "鏈瘑鍒埌涓婁笅鑱攖runk閾捐矾淇℃伅",
            "uplink_plan": [],
            "logs": logs or [],
        }

    worker_logs: list[dict] = [] if logs is None else logs

    def fetch_core():
        return _fetch_interface_config(core_switch, core_uplink, 20, 120, worker_logs)

    def fetch_access():
        return _fetch_interface_config(access_switch, access_uplink, 20, 120, worker_logs)

    with ThreadPoolExecutor(max_workers=2) as executor:
        core_future = executor.submit(fetch_core)
        access_future = executor.submit(fetch_access)
        core_uplink_output = core_future.result()
        access_uplink_output = access_future.result()

    uplink_plan = _build_exact_uplink_checks_dual(core_uplink, access_uplink, core_uplink_output, access_uplink_output, new_vlan)
    core_has_vlan = bool(uplink_plan[0]["has_vlan"]) if len(uplink_plan) > 0 else False
    access_has_vlan = bool(uplink_plan[1]["has_vlan"]) if len(uplink_plan) > 1 else False
    can_modify = core_has_vlan and access_has_vlan
    reason = "" if can_modify else "目标VLAN未在上下联trunk放通"

    result = {
        "core_uplink_interface": core_uplink,
        "access_uplink_interface": access_uplink,
        "core_allowed_vlans": uplink_plan[0]["allowed_vlans"] if len(uplink_plan) > 0 else "",
        "access_allowed_vlans": uplink_plan[1]["allowed_vlans"] if len(uplink_plan) > 1 else "",
        "core_has_vlan": core_has_vlan,
        "access_has_vlan": access_has_vlan,
        "can_modify": can_modify,
        "reason": reason,
        "uplink_plan": uplink_plan,
        "logs": worker_logs,
    }
    _cache_set(cache_key, result, 90)
    return result


def _vlan_in_spec(spec_text: str, vlan_id: int) -> bool:
    tokens = re.split(r"\s+", str(spec_text or "").strip())
    idx = 0
    while idx < len(tokens):
        token = tokens[idx]
        if token in {"to", ""}:
            idx += 1
            continue
        if token == "all":
            return True
        if token.isdigit():
            if int(token) == vlan_id:
                return True
            if idx + 2 < len(tokens) and tokens[idx + 1] == "to" and tokens[idx + 2].isdigit():
                start = int(token)
                end = int(tokens[idx + 2])
                if start <= vlan_id <= end:
                    return True
                idx += 3
                continue
        idx += 1
    return False


def _find_uplink_candidates(config_text: str) -> list[dict]:
    candidates = []
    for iface, lines in _parse_interface_blocks(config_text).items():
        block_text = "\n".join(lines)
        lower = block_text.lower()
        desc_match = re.search(r"^\s*description\s+(.+)$", block_text, re.IGNORECASE | re.MULTILINE)
        description = desc_match.group(1).strip() if desc_match else ""

        port_type = ""
        if "port link-type trunk" in lower:
            port_type = "trunk"
        elif "port link-type hybrid" in lower:
            port_type = "hybrid"

        is_uplink = False
        if iface.lower().startswith("eth-trunk"):
            is_uplink = True
        if any(key in lower for key in ["uplink", "涓婅仈", "core", "鏍稿績", "姹囪仛"]):
            is_uplink = True
        if port_type in {"trunk", "hybrid"} and description:
            is_uplink = True

        if is_uplink and port_type in {"trunk", "hybrid"}:
            candidates.append(
                {
                    "interface_name": iface,
                    "port_type": port_type,
                    "description": description,
                    "already_allows": False,
                    "requires_add": False,
                }
            )
    return candidates


def _build_uplink_add_command(interface_name: str, port_type: str, new_vlan: int) -> list[str]:
    if port_type == "trunk":
        vlan_cmd = f"port trunk allow-pass vlan {new_vlan}"
    else:
        vlan_cmd = f"port hybrid tagged vlan {new_vlan}"
    return [f"interface {interface_name}", vlan_cmd, "quit"]


def _locate_with_logs(db: Session, target_ip: str) -> tuple[dict, list[dict]]:
    normalized_ip = _validate_ip(target_ip)
    logs: list[dict] = []

    core = _select_core_switch(db)
    if not core:
        raise ValueError("鏈壘鍒版牳蹇冧氦鎹㈡満")

    try:
        conn = _connect_device(core, timeout=20)
    except Exception as exc:
        raise ValueError(f"鏍稿績浜ゆ崲鏈鸿繛鎺ュけ璐? {exc}") from exc

    try:
        arp_cmd = f"display arp | include {normalized_ip}"
        arp_output = _run_show(conn, arp_cmd, 20)
        logs.append({"device": core.name, "command": arp_cmd, "output": arp_output[-1500:]})

        if not _parse_arp_mac(arp_output):
            arp_full_cmd = "display arp"
            arp_full_output = _run_show(conn, arp_full_cmd, 30)
            arp_output = _extract_arp_section(arp_full_output, normalized_ip)
            logs.append({"device": core.name, "command": arp_full_cmd, "output": arp_output[-1500:] if arp_output else "鏈粠瀹屾暣ARP琛ㄤ腑鍖归厤鍒扮洰鏍嘔P"})
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass

    mac = _parse_arp_mac(arp_output)
    if not mac:
        raise ValueError("ARP查不到")

    core_uplink = _parse_arp_interface(arp_output)
    core_uplink_desc = ""
    core_uplink_output = ""
    topology_row = _find_topology_row(core_uplink)
    topology_switch = _find_access_switch_by_row(db, topology_row)
    hinted_switch = None
    if topology_row:
        logs.append(
            {
                "device": "system",
                "command": "topology-map",
                "output": json.dumps(
                    {
                        "core_uplink_interface": core_uplink or "",
                        "topology_name": topology_row.get("name", ""),
                        "topology_ip": topology_row.get("ip", ""),
                        "topology_eth_trunk": topology_row.get("eth_trunk", ""),
                        "topology_neighbor_intf": topology_row.get("neighbor_intf", ""),
                        "matched_device": getattr(topology_switch, "name", ""),
                    },
                    ensure_ascii=False,
                ),
            }
        )
    if core_uplink:
        try:
            conn = _connect_device(core, timeout=15)
        except Exception as exc:
            logs.append({"device": core.name, "command": "connect", "output": f"reconnect failed: {exc}"})
        else:
            try:
                _core_uplink_cmd, core_uplink_output = _run_interface_show(
                    conn,
                    "display current-configuration interface",
                    core_uplink,
                    20,
                    logs,
                    core.name,
                )
                core_uplink_desc = _extract_description_name(core_uplink_output)
                hinted_switch = _find_access_switch_by_hint(db, core_uplink_desc)
            finally:
                try:
                    conn.disconnect()
                except Exception:
                    pass

    matches = []
    candidate_switches = [hinted_switch] if hinted_switch else _list_access_switches(db, normalized_ip)
    for switch in [item for item in candidate_switches if item]:
        try:
            conn = _connect_device(switch, timeout=20)
        except Exception as exc:
            logs.append({"device": switch.name, "command": "connect", "output": f"SSH connect failed: {exc}"})
            continue
        try:
            for mac_cmd in _build_mac_lookup_commands(mac):
                mac_output = _run_show(conn, mac_cmd, 12)
                logs.append({"device": switch.name, "command": mac_cmd, "output": mac_output[-1500:]})
                interface_name = _parse_port_from_mac(mac_output)
                if interface_name:
                    matches.append({"switch": switch, "interface_name": interface_name, "mac_output": mac_output})
                    break
        finally:
            try:
                conn.disconnect()
            except Exception:
                pass

    if not matches:
        raise ValueError("MAC查不到")
    if len(matches) > 1:
        names = ", ".join(f"{item['switch'].name}:{item['interface_name']}" for item in matches[:5])
        raise ValueError(f"鎵惧埌澶氫釜绔彛锛屾棤娉曞敮涓€瀹氫綅: {names}")

    target = matches[0]
    access_switch: Device = target["switch"]
    interface_name = target["interface_name"]

    try:
        conn = _connect_device(access_switch, timeout=20)
    except Exception as exc:
        raise ValueError(f"鎺ュ叆浜ゆ崲鏈鸿繛鎺ュけ璐? {exc}") from exc

    try:
        _config_cmd, config_output = _run_interface_show(
            conn,
            "display current-configuration interface",
            interface_name,
            20,
            logs,
            access_switch.name,
        )

        _detail_cmd, detail_output = _run_interface_show(
            conn,
            "display interface",
            interface_name,
            20,
            logs,
            access_switch.name,
        )

        if core_uplink:
            _access_uplink_cmd, _access_uplink_output = _run_interface_show(
                conn,
                "display current-configuration interface",
                core_uplink,
                20,
                logs,
                access_switch.name,
            )

        switch_cfg_cmd = "display current-configuration"
        switch_cfg_output = _run_show(conn, switch_cfg_cmd, 40)
        logs.append({"device": access_switch.name, "command": switch_cfg_cmd, "output": switch_cfg_output[-3000:]})
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass

    profile = _parse_port_profile(config_output)
    forbidden, forbidden_reason = _is_forbidden_port(interface_name, profile)
    if forbidden:
        raise ValueError(forbidden_reason)
    if profile["port_type"] != "access":
        raise ValueError("当前端口不是 access 口")

    uplinks = _find_uplink_candidates(switch_cfg_output)

    result = {
        "target_ip": normalized_ip,
        "mac": mac,
        "core_switch": {"id": core.id, "name": core.name, "ip": core.ip},
        "access_switch": {"id": access_switch.id, "name": access_switch.name, "ip": access_switch.ip},
        "core_uplink_interface": core_uplink or "",
        "core_uplink_description": core_uplink_desc,
        "interface_name": interface_name,
        "current_vlan": profile["current_vlan"],
        "port_status": _parse_port_status(detail_output),
        "port_type": profile["port_type"],
        "port_description": profile["description"],
        "uplink_candidates": uplinks,
    }
    return result, logs


def locate_ip(db: Session, target_ip: str) -> dict:
    result, logs = _locate_with_logs(db, target_ip)
    return {**result, "logs": logs}


def _build_precheck(db: Session, locate_data: dict, new_vlan: int) -> dict:
    if not isinstance(new_vlan, int) or not (1 <= new_vlan <= 4094):
        raise ValueError("新VLAN不合法")

    access_switch = locate_data.get("access_switch") or {}
    switch_id = int(access_switch.get("id") or 0)
    interface_name = str(locate_data.get("interface_name") or "").strip()
    old_vlan = str(locate_data.get("current_vlan") or "").strip()
    port_type = str(locate_data.get("port_type") or "").strip().lower()
    if not switch_id or not interface_name:
        raise ValueError("定位结果无效，缺少交换机或端口信息")
    if str(new_vlan) == old_vlan:
        raise ValueError("鏂癡LAN涓庢棫VLAN鐩稿悓")
    if port_type != "access":
        raise ValueError("当前端口不是 access 口")

    switch = db.query(Device).filter(Device.id == switch_id).first()
    if not switch:
        raise ValueError("鎺ュ叆浜ゆ崲鏈轰笉瀛樺湪")

    logs: list[dict] = []
    try:
        conn = _connect_device(switch, timeout=20)
    except Exception as exc:
        raise ValueError(f"鎺ュ叆浜ゆ崲鏈鸿繛鎺ュけ璐? {exc}") from exc

    try:
        vlan_cmd = f"display vlan {new_vlan}"
        vlan_output = _run_show(conn, vlan_cmd, 20)
        logs.append({"device": switch.name, "command": vlan_cmd, "output": vlan_output[-1200:]})
        if not re.search(rf"\b{new_vlan}\b", vlan_output):
            raise ValueError("目标VLAN不存在")

        interface_cmd = f"display current-configuration interface {interface_name}"
        interface_output = _run_show(conn, interface_cmd, 20)
        logs.append({"device": switch.name, "command": interface_cmd, "output": interface_output[-2000:]})

        profile = _parse_port_profile(interface_output)
        forbidden, reason = _is_forbidden_port(interface_name, profile)
        if forbidden:
            raise ValueError(reason)
        if profile["port_type"] != "access":
            raise ValueError("当前端口不是 access 口")

        core_uplink = str(locate_data.get("core_uplink_interface") or "").strip()
        access_uplink_output = ""
        if core_uplink:
            _access_uplink_cmd, access_uplink_output = _run_interface_show(
                conn,
                "display current-configuration interface",
                core_uplink,
                20,
                logs,
                switch.name,
            )
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass

    uplink_plan = []
    warnings = []
    core_uplink_output = ""
    core_uplink = str(locate_data.get("core_uplink_interface") or "").strip()
    core_switch_data = locate_data.get("core_switch") or {}
    core_switch_id = int(core_switch_data.get("id") or 0)
    if core_uplink and core_switch_id:
        core_switch = db.query(Device).filter(Device.id == core_switch_id).first()
        if core_switch:
            try:
                conn = _connect_device(core_switch, timeout=20)
            except Exception as exc:
                warnings.append(f"鏍稿績浜ゆ崲鏈轰笂鑱斿彛妫€鏌ュけ璐? {exc}")
            else:
                try:
                    _core_uplink_cmd, core_uplink_output = _run_interface_show(
                        conn,
                        "display current-configuration interface",
                        core_uplink,
                        20,
                        logs,
                        core_switch.name,
                    )
                finally:
                    try:
                        conn.disconnect()
                    except Exception:
                        pass
    uplink_plan = _build_exact_uplink_checks_dual(core_uplink, access_uplink, core_uplink_output, access_uplink_output, new_vlan)

    if not uplink_plan:
        warnings.append("鏈瘑鍒埌鏄庣‘鐨勪笂鑱?涓嬭仈鏀鹃€氱鍙ｏ紝璇蜂汉宸ュ鏍搁摼璺疺LAN閫忎紶鎯呭喌")

    core_has_vlan = bool(uplink_plan[0]["has_vlan"]) if len(uplink_plan) > 0 else False
    access_has_vlan = bool(uplink_plan[1]["has_vlan"]) if len(uplink_plan) > 1 else False
    can_modify = core_has_vlan and access_has_vlan
    reason = "" if can_modify else "目标VLAN未在上下联trunk放通"
    if not can_modify:
        raise ValueError(reason)

    access_commands = [f"interface {interface_name}", f"port default vlan {new_vlan}", "quit"]
    return {
        "target_ip": locate_data.get("target_ip", ""),
        "mac": locate_data.get("mac", ""),
        "access_switch": access_switch,
        "interface_name": interface_name,
        "old_vlan": old_vlan,
        "new_vlan": str(new_vlan),
        "core_allowed_vlans": uplink_plan[0]["allowed_vlans"] if len(uplink_plan) > 0 else "",
        "access_allowed_vlans": uplink_plan[1]["allowed_vlans"] if len(uplink_plan) > 1 else "",
        "core_has_vlan": core_has_vlan,
        "access_has_vlan": access_has_vlan,
        "can_modify": can_modify,
        "reason": reason,
        "access_commands": access_commands,
        "uplink_plan": uplink_plan,
        "warnings": warnings,
        "logs": logs,
    }


def execute_vlan_change(db: Session, locate_data: dict, new_vlan: int, operator: str = "unknown") -> dict:
    precheck = _build_precheck(db, locate_data, new_vlan)

    switch_id = int((locate_data.get("access_switch") or {}).get("id") or 0)
    switch = db.query(Device).filter(Device.id == switch_id).first()
    if not switch:
        raise ValueError("鎺ュ叆浜ゆ崲鏈轰笉瀛樺湪")

    executed_steps = []
    verify_steps = []
    target_ip = str(locate_data.get("target_ip") or "").strip()
    target_mac = _normalize_mac(str(locate_data.get("mac") or "").strip())
    interface_name = str(locate_data.get("interface_name") or "").strip()
    old_vlan = str(locate_data.get("current_vlan") or "").strip()

    status = "failed"
    message = ""
    verify_result = ""
    vlan_change_success = False
    shutdown_success = False
    undo_shutdown_success = False
    refresh_attempted = False
    core_uplink_added = False
    access_uplink_added = False
    core_uplink_added = False
    access_uplink_added = False

    try:
        conn = _connect_device(switch, timeout=20)
    except Exception as exc:
        raise ValueError(f"鎺ュ叆浜ゆ崲鏈鸿繛鎺ュけ璐? {exc}") from exc

    try:
        uplink_additions = [item for item in precheck["uplink_plan"] if item["requires_add"]]
        for uplink in uplink_additions:
            commands = uplink["planned_commands"]
            conn.send_config_set(commands, read_timeout=40)
            executed_steps.append(
                {
                    "device": switch.name,
                    "stage": "uplink_allow",
                    "interface_name": uplink["interface_name"],
                    "commands": commands,
                }
            )

        conn.send_config_set(precheck["access_commands"], read_timeout=40)
        executed_steps.append(
            {
                "device": switch.name,
                "stage": "access_change",
                "interface_name": interface_name,
                "commands": precheck["access_commands"],
            }
        )
        vlan_change_success = True

        # Some devices do not accept short interface names (e.g. GE0/0/3) for show commands.
        # Try multiple variants and record every attempt into verify_steps.
        verify_access_cmd, verify_access_output = _run_interface_show(
            conn,
            "display current-configuration interface",
            interface_name,
            20,
            verify_steps,
            switch.name,
        )
        verify_profile = _parse_port_profile(verify_access_output)
        if str(verify_profile["current_vlan"]) != str(new_vlan):
            raise ValueError(f"修改后校验失败: 期望VLAN={new_vlan} 实际VLAN={verify_profile['current_vlan']} (命令: {verify_access_cmd})")

        for uplink in uplink_additions:
            verify_uplink_cmd = f"display current-configuration interface {uplink['interface_name']}"
            verify_uplink_output = _run_show(conn, verify_uplink_cmd, 20)
            verify_steps.append({"device": switch.name, "command": verify_uplink_cmd, "output": verify_uplink_output[-2000:]})
            if not _interface_vlan_allowed(verify_uplink_output, new_vlan):
                raise ValueError(f"涓婅仈/涓嬭仈绔彛 {uplink['interface_name']} 鏈斁閫?VLAN {new_vlan}")

        status = "success"
        message = "VLAN淇敼鎴愬姛骞舵牎楠岄€氳繃"
        verify_result = json.dumps({"precheck": precheck, "verify_steps": verify_steps}, ensure_ascii=False)
    except Exception as exc:
        message = str(exc)
        verify_result = json.dumps({"precheck": precheck, "verify_steps": verify_steps, "error": message}, ensure_ascii=False)
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass

    log_row = IpVlanChangeLog(
        operator=str(operator or "unknown"),
        target_ip=target_ip,
        target_mac=target_mac,
        core_switch=str((locate_data.get("core_switch") or {}).get("name") or ""),
        access_switch=str((locate_data.get("access_switch") or {}).get("name") or ""),
        interface_name=interface_name,
        port_type=str(locate_data.get("port_type") or ""),
        port_status=str(locate_data.get("port_status") or ""),
        port_description=str(locate_data.get("port_description") or ""),
        old_vlan=old_vlan,
        new_vlan=str(new_vlan),
        status=status,
        message=message,
        commands=json.dumps({"executed_steps": executed_steps, "precheck": precheck}, ensure_ascii=False),
        verify_result=verify_result,
        execute_time=_now_iso(),
    )
    db.add(log_row)
    db.commit()
    db.refresh(log_row)

    if status != "success":
        raise ValueError(message)

    return {
        "id": log_row.id,
        "status": status,
        "message": message,
        "target_ip": target_ip,
        "target_mac": target_mac,
        "access_switch": locate_data.get("access_switch", {}),
        "interface_name": interface_name,
        "old_vlan": old_vlan,
        "new_vlan": str(new_vlan),
        "operator": log_row.operator,
        "execute_time": log_row.execute_time,
        "warnings": precheck.get("warnings", []),
    }


def _locate_with_logs_v2(db: Session, target_ip: str) -> tuple[dict, list[dict]]:
    normalized_ip = _validate_ip(target_ip)
    logs: list[dict] = []
    cached_result = _cache_get(f"locate:{normalized_ip}")
    if cached_result is not None:
        logs.append({"device": "system", "command": "locate-cache", "output": f"鍛戒腑缂撳瓨 {normalized_ip}"})
        return dict(cached_result), logs

    core = _select_core_switch(db)
    if not core:
        raise LocateTraceError("鏈壘鍒版牳蹇冧氦鎹㈡満", logs)

    try:
        conn = _connect_device(core, timeout=20)
    except Exception as exc:
        raise LocateTraceError(f"鏍稿績浜ゆ崲鏈鸿繛鎺ュけ璐? {exc}", logs) from exc

    try:
        logs.append({"device": core.name, "command": "ssh-connect", "output": f"SSH鐧诲綍鎴愬姛 {core.ip}"})
        arp_cmd = f"display arp | include {normalized_ip}"
        arp_output = _run_show(conn, arp_cmd, 20)
        logs.append({"device": core.name, "command": arp_cmd, "output": arp_output[-1500:]})
        if not _parse_arp_mac(arp_output):
            arp_full_cmd = "display arp"
            arp_full_output = _run_show(conn, arp_full_cmd, 30)
            arp_output = _extract_arp_section(arp_full_output, normalized_ip)
            logs.append(
                {
                    "device": core.name,
                    "command": arp_full_cmd,
                    "output": arp_output[-1500:] if arp_output else "鏈粠瀹屾暣ARP琛ㄤ腑鍖归厤鍒扮洰鏍嘔P",
                }
            )
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass

    mac = _parse_arp_mac(arp_output)
    if not mac:
        raise LocateTraceError("ARP查不到", logs)

    core_uplink = _parse_arp_interface(arp_output)
    core_uplink_desc = ""
    topology_row = _find_topology_row(core_uplink)
    topology_switch = _find_access_switch_by_row(db, topology_row)
    if topology_row:
        logs.append(
            {
                "device": "system",
                "command": "topology-map",
                "output": json.dumps(
                    {
                        "core_uplink_interface": core_uplink or "",
                        "topology_name": topology_row.get("name", ""),
                        "topology_ip": topology_row.get("ip", ""),
                        "topology_eth_trunk": topology_row.get("eth_trunk", ""),
                        "topology_neighbor_intf": topology_row.get("neighbor_intf", ""),
                        "matched_device": getattr(topology_switch, "name", ""),
                    },
                    ensure_ascii=False,
                ),
            }
        )
    hinted_switch = None
    if core_uplink:
        try:
            conn = _connect_device(core, timeout=15)
        except Exception as exc:
            logs.append({"device": core.name, "command": "connect", "output": f"reconnect failed: {exc}"})
        else:
            try:
                _cmd, core_uplink_output = _run_interface_show(
                    conn,
                    "display current-configuration interface",
                    core_uplink,
                    20,
                    logs,
                    core.name,
                )
                core_uplink_desc = _extract_description_name(core_uplink_output)
                if not topology_row:
                    topology_row = _find_topology_row(core_uplink, core_uplink_desc)
                    topology_switch = _find_access_switch_by_row(db, topology_row)
                hinted_switch = topology_switch or _find_access_switch_by_hint(db, core_uplink_desc)
            finally:
                try:
                    conn.disconnect()
                except Exception:
                    pass

    matches = []
    candidate_switches = _ordered_access_switches(db, normalized_ip, topology_switch or hinted_switch)
    for switch in [item for item in candidate_switches if item]:
        try:
            conn = _connect_device(switch, timeout=30)
        except Exception as exc:
            logs.append({"device": switch.name, "command": "connect", "output": f"SSH connect failed: {exc}"})
            if len(candidate_switches) == 1:
                raise LocateTraceError(f"鎺ュ叆浜ゆ崲鏈鸿繛鎺ュけ璐? {switch.name} {switch.ip} - {exc}", logs) from exc
            continue
        try:
            logs.append({"device": switch.name, "command": "ssh-connect", "output": f"SSH鐧诲綍鎴愬姛 {switch.ip}"})
            for mac_cmd in _build_mac_lookup_commands(mac):
                mac_output = _run_show(conn, mac_cmd, 12)
                logs.append({"device": switch.name, "command": mac_cmd, "output": mac_output[-1500:]})
                interface_name = _parse_port_from_mac(mac_output)
                if interface_name:
                    matches.append({"switch": switch, "interface_name": interface_name, "mac_output": mac_output})
                    break
        finally:
            try:
                conn.disconnect()
            except Exception:
                pass

    if not matches:
        summary = {
            "target_ip": normalized_ip,
            "mac": mac,
            "core_uplink": core_uplink or "",
            "core_uplink_description": core_uplink_desc,
            "candidate_switches": [item.name for item in candidate_switches if item],
        }
        logs.append({"device": "system", "command": "locate-summary", "output": json.dumps(summary, ensure_ascii=False)})
        raise LocateTraceError("MAC查不到", logs)
    if len(matches) > 1:
        names = ", ".join(f"{item['switch'].name}:{item['interface_name']}" for item in matches[:5])
        raise LocateTraceError(f"鎵惧埌澶氫釜绔彛锛屾棤娉曞敮涓€瀹氫綅: {names}", logs)

    target = matches[0]
    access_switch: Device = target["switch"]
    interface_name = target["interface_name"]

    try:
        conn = _connect_device(access_switch, timeout=20)
    except Exception as exc:
        raise LocateTraceError(f"鎺ュ叆浜ゆ崲鏈鸿繛鎺ュけ璐? {exc}", logs) from exc

    try:
        _cfg_cmd, config_output = _run_interface_show(
            conn,
            "display current-configuration interface",
            interface_name,
            20,
            logs,
            access_switch.name,
        )
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass

    profile = _parse_port_profile(config_output)
    forbidden, forbidden_reason = _is_forbidden_port(interface_name, profile)
    if forbidden:
        raise LocateTraceError(forbidden_reason, logs)
    if profile["port_type"] != "access":
        raise LocateTraceError("当前端口不是 access 口", logs)

    result = {
        "target_ip": normalized_ip,
        "mac": mac,
        "core_switch": {"id": core.id, "name": core.name, "ip": core.ip},
        "access_switch": {"id": access_switch.id, "name": access_switch.name, "ip": access_switch.ip},
        "core_uplink_interface": core_uplink or "",
        "interface_name": interface_name,
        "current_vlan": profile["current_vlan"],
        "core_allowed_vlans": uplink_snapshot.get("core_allowed_vlans", ""),
        "access_allowed_vlans": uplink_snapshot.get("access_allowed_vlans", ""),
        "core_has_vlan": False,
        "access_has_vlan": False,
        "can_modify": False,
        "uplink_plan": uplink_snapshot.get("uplink_plan", []),
        "reason": "" if (core_has_vlan and access_has_vlan) else "目标VLAN未在上下联trunk放通",
        "uplink_candidates": uplinks,
    }, logs


def locate_ip(db: Session, target_ip: str) -> dict:
    result, logs = _locate_with_logs_v2(db, target_ip)
    return {**result, "logs": logs}


def _locate_basic_with_logs(db: Session, target_ip: str) -> tuple[dict, list[dict]]:
    normalized_ip = _validate_ip(target_ip)
    logs: list[dict] = []

    cached_result = _cache_get(f"locate:{normalized_ip}")
    if cached_result is not None:
        logs.append({"device": "system", "command": "locate-cache", "output": f"鍛戒腑缂撳瓨 {normalized_ip}"})
        return dict(cached_result), logs

    core = _select_core_switch(db)
    if not core:
        raise LocateTraceError("鏈壘鍒版牳蹇冧氦鎹㈡満", logs)

    arp_cache_key = f"arp:{normalized_ip}"
    arp_data = _cache_get(arp_cache_key)
    if arp_data is None:
        try:
            conn = _connect_device(core, timeout=20)
        except Exception as exc:
            raise LocateTraceError(f"鏍稿績浜ゆ崲鏈鸿繛鎺ュけ璐? {exc}", logs) from exc

        try:
            logs.append({"device": core.name, "command": "ssh-connect", "output": f"SSH鐧诲綍鎴愬姛 {core.ip}"})
            arp_cmd = f"display arp | include {normalized_ip}"
            arp_output = _run_show(conn, arp_cmd, 20)
            logs.append({"device": core.name, "command": arp_cmd, "output": arp_output[-1500:]})
            if not _parse_arp_mac(arp_output):
                arp_full_cmd = "display arp"
                arp_full_output = _run_show(conn, arp_full_cmd, 30)
                arp_output = _extract_arp_section(arp_full_output, normalized_ip)
                logs.append(
                    {
                        "device": core.name,
                        "command": arp_full_cmd,
                        "output": arp_output[-1500:] if arp_output else "鏈粠瀹屾暣ARP琛ㄤ腑鍖归厤鍒扮洰鏍嘔P",
                    }
                )
        finally:
            try:
                conn.disconnect()
            except Exception:
                pass

        arp_data = {
            "mac": _parse_arp_mac(arp_output),
            "core_uplink": _parse_arp_interface(arp_output) or "",
        }
        _cache_set(arp_cache_key, arp_data, 45)
    else:
        logs.append({"device": core.name, "command": "arp-cache", "output": json.dumps(arp_data, ensure_ascii=False)})

    mac = str(arp_data.get("mac") or "").strip()
    if not mac:
        raise LocateTraceError("ARP查不到", logs)

    core_uplink = str(arp_data.get("core_uplink") or "").strip()
    core_uplink_desc = ""
    topology_row = _find_topology_row(core_uplink)
    topology_switch = _find_access_switch_by_row(db, topology_row)
    if topology_row:
        logs.append(
            {
                "device": "system",
                "command": "topology-map",
                "output": json.dumps(
                    {
                        "core_uplink_interface": core_uplink,
                        "topology_name": topology_row.get("name", ""),
                        "topology_ip": topology_row.get("ip", ""),
                        "topology_eth_trunk": topology_row.get("eth_trunk", ""),
                        "topology_neighbor_intf": topology_row.get("neighbor_intf", ""),
                        "matched_device": getattr(topology_switch, "name", ""),
                    },
                    ensure_ascii=False,
                ),
            }
        )

    hinted_switch = None
    if core_uplink and not topology_switch:
        try:
            core_uplink_output = _fetch_interface_config(core, core_uplink, 20, 120, logs)
            core_uplink_desc = _extract_description_name(core_uplink_output)
        except Exception as exc:
            logs.append({"device": core.name, "command": "core-uplink-check", "output": f"failed: {exc}"})
            core_uplink_output = ""
        if not topology_row:
            topology_row = _find_topology_row(core_uplink, core_uplink_desc)
            topology_switch = _find_access_switch_by_row(db, topology_row)
        hinted_switch = topology_switch or _find_access_switch_by_hint(db, core_uplink_desc)
        logs.append(
            {
                "device": "system",
                "command": "core-uplink-map",
                "output": json.dumps(
                    {
                        "core_uplink_interface": core_uplink,
                        "core_uplink_description": core_uplink_desc,
                        "topology_name": (topology_row or {}).get("name", ""),
                        "topology_ip": (topology_row or {}).get("ip", ""),
                        "hinted_switch": getattr(hinted_switch, "name", ""),
                    },
                    ensure_ascii=False,
                ),
            }
        )

    matches = []
    candidate_switches = _ordered_access_switches(db, normalized_ip, topology_switch or hinted_switch)
    for switch in [item for item in candidate_switches if item]:
        try:
            conn = _connect_device(switch, timeout=30)
        except Exception as exc:
            logs.append({"device": switch.name, "command": "connect", "output": f"SSH connect failed: {exc}"})
            if len(candidate_switches) == 1:
                raise LocateTraceError(f"鎺ュ叆浜ゆ崲鏈鸿繛鎺ュけ璐? {switch.name} {switch.ip} - {exc}", logs) from exc
            continue
        try:
            logs.append({"device": switch.name, "command": "ssh-connect", "output": f"SSH鐧诲綍鎴愬姛 {switch.ip}"})
            for mac_cmd in _build_mac_lookup_commands(mac):
                mac_output = _run_show(conn, mac_cmd, 12)
                logs.append({"device": switch.name, "command": mac_cmd, "output": mac_output[-1500:]})
                interface_name = _parse_port_from_mac(mac_output)
                if interface_name:
                    matches.append({"switch": switch, "interface_name": interface_name})
                    break
        finally:
            try:
                conn.disconnect()
            except Exception:
                pass

    if not matches:
        logs.append(
            {
                "device": "system",
                "command": "locate-summary",
                "output": json.dumps(
                    {
                        "target_ip": normalized_ip,
                        "mac": mac,
                        "core_uplink": core_uplink,
                        "core_uplink_description": core_uplink_desc,
                        "candidate_switches": [item.name for item in candidate_switches if item],
                    },
                    ensure_ascii=False,
                ),
            }
        )
        raise LocateTraceError("MAC查不到", logs)

    if len(matches) > 1:
        names = ", ".join(f"{item['switch'].name}:{item['interface_name']}" for item in matches[:5])
        raise LocateTraceError(f"鎵惧埌澶氫釜绔彛锛屾棤娉曞敮涓€瀹氫綅: {names}", logs)

    target = matches[0]
    access_switch: Device = target["switch"]
    interface_name = target["interface_name"]

    try:
        config_output = _fetch_interface_config(access_switch, interface_name, 20, 60, logs)
    except Exception as exc:
        raise LocateTraceError(f"鎺ュ叆浜ゆ崲鏈鸿繛鎺ュけ璐? {exc}", logs) from exc

    profile = _parse_port_profile(config_output)
    forbidden, forbidden_reason = _is_forbidden_port(interface_name, profile)
    if forbidden:
        raise LocateTraceError(forbidden_reason, logs)
    if profile["port_type"] != "access":
        raise LocateTraceError("当前端口不是 access 口", logs)

    uplink_snapshot = _fetch_uplink_snapshot(
        db,
        {
            "core_switch": {"id": core.id, "name": core.name, "ip": core.ip},
            "access_switch": {"id": access_switch.id, "name": access_switch.name, "ip": access_switch.ip},
            "core_uplink_interface": core_uplink,
        },
        logs,
    )

    result = {
        "target_ip": normalized_ip,
        "mac": mac,
        "core_switch": {"id": core.id, "name": core.name, "ip": core.ip},
        "access_switch": {"id": access_switch.id, "name": access_switch.name, "ip": access_switch.ip},
        "core_uplink_interface": core_uplink,
        "access_uplink_interface": uplink_snapshot.get("access_uplink_interface", core_uplink),
        "interface_name": interface_name,
        "current_vlan": profile["current_vlan"],
        "core_allowed_vlans": uplink_snapshot.get("core_allowed_vlans", ""),
        "access_allowed_vlans": uplink_snapshot.get("access_allowed_vlans", ""),
        "core_has_vlan": False,
        "access_has_vlan": False,
        "can_modify": False,
        "uplink_plan": uplink_snapshot.get("uplink_plan", []),
        "reason": "待输入新VLAN后校验",
        "uplink_candidates": [],
    }
    _cache_set(f"locate:{normalized_ip}", result, 60)
    return result, logs


def _locate_from_mac_and_uplink(db: Session, mac: str, core_uplink: str, logs: list[dict], target_ip: str = "") -> dict:
    core = _select_core_switch(db)
    if not core:
        raise LocateTraceError("鏈壘鍒版牳蹇冧氦鎹㈡満", logs)

    core_uplink_desc = ""
    topology_row = _find_topology_row(core_uplink)
    topology_switch = _find_access_switch_by_row(db, topology_row)
    if topology_row:
        logs.append(
            {
                "device": "system",
                "command": "topology-map",
                "output": json.dumps(
                    {
                        "core_uplink_interface": core_uplink,
                        "topology_name": topology_row.get("name", ""),
                        "topology_ip": topology_row.get("ip", ""),
                        "topology_eth_trunk": topology_row.get("eth_trunk", ""),
                        "topology_neighbor_intf": topology_row.get("neighbor_intf", ""),
                        "matched_device": getattr(topology_switch, "name", ""),
                    },
                    ensure_ascii=False,
                ),
            }
        )

    hinted_switch = None
    if core_uplink and not topology_switch:
        try:
            core_uplink_output = _fetch_interface_config(core, core_uplink, 20, 120, logs)
            core_uplink_desc = _extract_description_name(core_uplink_output)
        except Exception as exc:
            logs.append({"device": core.name, "command": "core-uplink-check", "output": f"failed: {exc}"})
            core_uplink_output = ""
        if not topology_row:
            topology_row = _find_topology_row(core_uplink, core_uplink_desc)
            topology_switch = _find_access_switch_by_row(db, topology_row)
        hinted_switch = topology_switch or _find_access_switch_by_hint(db, core_uplink_desc)

    candidate_switches = _ordered_access_switches(db, target_ip, topology_switch or hinted_switch)
    matches = []
    for switch in [item for item in candidate_switches if item]:
        try:
            conn = _connect_device(switch, timeout=30)
        except Exception as exc:
            logs.append({"device": switch.name, "command": "connect", "output": f"SSH connect failed: {exc}"})
            if len(candidate_switches) == 1:
                raise LocateTraceError(f"鎺ュ叆浜ゆ崲鏈鸿繛鎺ュけ璐? {switch.name} {switch.ip} - {exc}", logs) from exc
            continue
        try:
            logs.append({"device": switch.name, "command": "ssh-connect", "output": f"SSH鐧诲綍鎴愬姛 {switch.ip}"})
            for mac_cmd in _build_mac_lookup_commands(mac):
                mac_output = _run_show(conn, mac_cmd, 12)
                logs.append({"device": switch.name, "command": mac_cmd, "output": mac_output[-1500:]})
                interface_name = _parse_port_from_mac(mac_output)
                if interface_name:
                    matches.append({"switch": switch, "interface_name": interface_name})
                    break
        finally:
            try:
                conn.disconnect()
            except Exception:
                pass

    if not matches:
        raise LocateTraceError("MAC查不到", logs)
    if len(matches) > 1:
        names = ", ".join(f"{item['switch'].name}:{item['interface_name']}" for item in matches[:5])
        raise LocateTraceError(f"鎵惧埌澶氫釜绔彛锛屾棤娉曞敮涓€瀹氫綅: {names}", logs)

    target = matches[0]
    access_switch: Device = target["switch"]
    interface_name = target["interface_name"]
    try:
        config_output = _fetch_interface_config(access_switch, interface_name, 20, 60, logs)
    except Exception as exc:
        raise LocateTraceError(f"鎺ュ叆浜ゆ崲鏈鸿繛鎺ュけ璐? {exc}", logs) from exc

    profile = _parse_port_profile(config_output)
    forbidden, forbidden_reason = _is_forbidden_port(interface_name, profile)
    if forbidden:
        raise LocateTraceError(forbidden_reason, logs)
    if profile["port_type"] != "access":
        raise LocateTraceError("当前端口不是 access 口", logs)

    return {
        "target_ip": str(target_ip or ""),
        "mac": mac,
        "core_switch": {"id": core.id, "name": core.name, "ip": core.ip},
        "access_switch": {"id": access_switch.id, "name": access_switch.name, "ip": access_switch.ip},
        "core_uplink_interface": core_uplink,
        "access_uplink_interface": uplink_snapshot.get("access_uplink_interface", core_uplink),
        "interface_name": interface_name,
        "current_vlan": profile["current_vlan"],
        "core_allowed_vlans": uplink_snapshot.get("core_allowed_vlans", ""),
        "access_allowed_vlans": uplink_snapshot.get("access_allowed_vlans", ""),
        "core_has_vlan": False,
        "access_has_vlan": False,
        "can_modify": False,
        "uplink_plan": uplink_snapshot.get("uplink_plan", []),
        "reason": "待输入新VLAN后校验",
        "uplink_candidates": [],
        "pending_selection": False,
        "match_options": [],
        "query_type": "mac" if not target_ip else "ip",
    }


def locate_by_mac(db: Session, mac_keyword: str, selected_mac: str = "", selected_core_uplink: str = "") -> dict:
    keyword = _normalize_mac_keyword(mac_keyword)
    if not keyword:
        raise ValueError("MAC关键字不能为空")

    logs: list[dict] = []
    core = _select_core_switch(db)
    if not core:
        raise LocateTraceError("鏈壘鍒版牳蹇冧氦鎹㈡満", logs)

    try:
        conn = _connect_device(core, timeout=20)
    except Exception as exc:
        raise LocateTraceError(f"鏍稿績浜ゆ崲鏈鸿繛鎺ュけ璐? {exc}", logs) from exc

    try:
        logs.append({"device": core.name, "command": "ssh-connect", "output": f"SSH鐧诲綍鎴愬姛 {core.ip}"})
        core_cmd = f"display mac-address | include {keyword}"
        core_output = _run_show(conn, core_cmd, 20)
        logs.append({"device": core.name, "command": core_cmd, "output": core_output[-2000:]})
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass

    core_matches = _extract_mac_matches(core_output)
    if not core_matches:
        raise LocateTraceError("鏍稿績浜ゆ崲鏈烘湭鍖归厤鍒癕AC", logs)

    normalized_selected = _normalize_mac(selected_mac) if selected_mac else ""
    if not normalized_selected:
        unique_options = []
        seen: set[tuple[str, str]] = set()
        for item in core_matches:
            interface_name = item.get("interface_name") or ""
            key = (item["mac"], interface_name)
            if key in seen:
                continue
            seen.add(key)
            unique_options.append(
                {
                    "mac": item["mac"],
                    "core_uplink_interface": interface_name,
                    "display_text": f"{item['mac']} / {interface_name or '-'}",
                }
            )
        if len(unique_options) > 1:
            return {
                "target_ip": "",
                "mac": "",
                "core_switch": {"id": core.id, "name": core.name, "ip": core.ip},
                "access_switch": None,
                "core_uplink_interface": "",
                "interface_name": "",
                "current_vlan": "",
                "core_allowed_vlans": "",
                "access_allowed_vlans": "",
                "core_has_vlan": False,
                "access_has_vlan": False,
                "can_modify": False,
                "reason": "鍖归厤鍒板涓狹AC锛岃鍏堥€夋嫨鍏蜂綋MAC",
                "uplink_candidates": [],
                "pending_selection": True,
                "match_options": unique_options,
                "query_type": "mac",
                "logs": logs,
            }
        normalized_selected = unique_options[0]["mac"]
        selected_core_uplink = unique_options[0]["core_uplink_interface"]

    chosen = next((item for item in core_matches if item["mac"] == normalized_selected), None)
    if not chosen:
        raise LocateTraceError("鏈壘鍒版墍閫塎AC", logs)

    core_uplink = str(selected_core_uplink or chosen.get("interface_name") or "").strip()
    result = _locate_from_mac_and_uplink(db, normalized_selected, core_uplink, logs, target_ip="")
    return {**result, "logs": logs}


def validate_target_vlan(db: Session, locate_data: dict, new_vlan: int) -> dict:
    if not isinstance(new_vlan, int) or not (1 <= new_vlan <= 4094):
        raise ValueError("鏂癡LAN鑼冨洿闇€鍦?~4094涔嬮棿")

    logs: list[dict] = []
    result = _validate_uplink_pass_through(db, locate_data, new_vlan, logs)
    return result


def locate_ip(db: Session, target_ip: str) -> dict:
    result, logs = _locate_basic_with_logs(db, target_ip)
    return {**result, "logs": logs}


def _build_precheck_v2(db: Session, locate_data: dict, new_vlan: int) -> dict:
    if not isinstance(new_vlan, int) or not (1 <= new_vlan <= 4094):
        raise ValueError("鏂癡LAN鑼冨洿闇€鍦?~4094涔嬮棿")

    access_switch = locate_data.get("access_switch") or {}
    switch_id = int(access_switch.get("id") or 0)
    interface_name = str(locate_data.get("interface_name") or "").strip()
    old_vlan = str(locate_data.get("current_vlan") or "").strip()
    if not switch_id or not interface_name:
        raise ValueError("定位结果无效，缺少交换机或端口信息")
    if str(new_vlan) == old_vlan:
        raise ValueError("鏂癡LAN涓庢棫VLAN鐩稿悓")

    switch = db.query(Device).filter(Device.id == switch_id).first()
    if not switch:
        raise ValueError("鎺ュ叆浜ゆ崲鏈轰笉瀛樺湪")

    logs: list[dict] = []
    try:
        conn = _connect_device(switch, timeout=20)
    except Exception as exc:
        raise ValueError(f"鎺ュ叆浜ゆ崲鏈鸿繛鎺ュけ璐? {exc}") from exc

    try:
        vlan_cmd = f"display vlan {new_vlan}"
        vlan_output = _run_show(conn, vlan_cmd, 20)
        logs.append({"device": switch.name, "command": vlan_cmd, "output": vlan_output[-1200:]})
        if not re.search(rf"\b{new_vlan}\b", vlan_output):
            raise ValueError("目标VLAN不存在")

        interface_output = _fetch_interface_config(switch, interface_name, 20, 60, logs)
        profile = _parse_port_profile(interface_output)
        forbidden, reason = _is_forbidden_port(interface_name, profile)
        if forbidden:
            raise ValueError(reason)
        if profile["port_type"] != "access":
            raise ValueError("当前端口不是 access 口")
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass

    validation = _validate_uplink_pass_through(db, locate_data, new_vlan, logs)

    interface_name_config = _normalize_interface_for_config(interface_name)
    access_commands = [f"interface {interface_name_config}", f"port default vlan {new_vlan}", "quit"]
    return {
        "target_ip": locate_data.get("target_ip", ""),
        "mac": locate_data.get("mac", ""),
        "access_switch": access_switch,
        "interface_name": interface_name,
        "interface_name_config": interface_name_config,
        "old_vlan": old_vlan,
        "new_vlan": str(new_vlan),
        "core_uplink_interface": validation.get("core_uplink_interface", locate_data.get("core_uplink_interface", "")),
        "access_uplink_interface": validation.get("access_uplink_interface", locate_data.get("access_uplink_interface", "")),
        "core_allowed_vlans": validation.get("core_allowed_vlans", ""),
        "access_allowed_vlans": validation.get("access_allowed_vlans", ""),
        "core_has_vlan": bool(validation.get("core_has_vlan")),
        "access_has_vlan": bool(validation.get("access_has_vlan")),
        "can_modify": True,
        "access_commands": access_commands,
        "uplink_plan": validation.get("uplink_plan", []),
        "warnings": [],
        "logs": logs,
    }


def execute_vlan_change(db: Session, locate_data: dict, new_vlan: int, operator: str = "unknown") -> dict:
    precheck = _build_precheck_v2(db, locate_data, new_vlan)

    switch_id = int((locate_data.get("access_switch") or {}).get("id") or 0)
    switch = db.query(Device).filter(Device.id == switch_id).first()
    if not switch:
        raise ValueError("鎺ュ叆浜ゆ崲鏈轰笉瀛樺湪")

    executed_steps = []
    verify_steps = []
    target_ip = str(locate_data.get("target_ip") or "").strip()
    target_mac = _normalize_mac(str(locate_data.get("mac") or "").strip())
    interface_name = str(locate_data.get("interface_name") or "").strip()
    old_vlan = str(locate_data.get("current_vlan") or "").strip()

    status = "failed"
    message = ""
    verify_result = ""

    try:
        conn = _connect_device(switch, timeout=20)
    except Exception as exc:
        raise ValueError(f"鎺ュ叆浜ゆ崲鏈鸿繛鎺ュけ璐? {exc}") from exc

    try:
        conn.send_config_set(precheck["access_commands"], read_timeout=40)
        executed_steps.append(
            {
                "device": switch.name,
                "stage": "access_change",
                "interface_name": interface_name,
                "commands": precheck["access_commands"],
            }
        )

        verify_access_cmd = f"display current-configuration interface {interface_name}"
        verify_access_output = _run_show(conn, verify_access_cmd, 20)
        verify_steps.append({"device": switch.name, "command": verify_access_cmd, "output": verify_access_output[-2000:]})
        verify_profile = _parse_port_profile(verify_access_output)
        if str(verify_profile["current_vlan"]) != str(new_vlan):
            raise ValueError("修改后校验失败")

        status = "success"
        message = "VLAN淇敼鎴愬姛骞舵牎楠岄€氳繃"
        verify_result = json.dumps({"precheck": precheck, "verify_steps": verify_steps}, ensure_ascii=False)
    except Exception as exc:
        message = str(exc)
        verify_result = json.dumps({"precheck": precheck, "verify_steps": verify_steps, "error": message}, ensure_ascii=False)
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass

    log_row = IpVlanChangeLog(
        operator=str(operator or "unknown"),
        target_ip=target_ip,
        target_mac=target_mac,
        core_switch=str((locate_data.get("core_switch") or {}).get("name") or ""),
        access_switch=str((locate_data.get("access_switch") or {}).get("name") or ""),
        interface_name=interface_name,
        port_type=str(locate_data.get("port_type") or ""),
        port_status=str(locate_data.get("port_status") or ""),
        port_description=str(locate_data.get("port_description") or ""),
        old_vlan=old_vlan,
        new_vlan=str(new_vlan),
        status=status,
        message=message,
        commands=json.dumps({"executed_steps": executed_steps, "precheck": precheck}, ensure_ascii=False),
        verify_result=verify_result,
        execute_time=_now_iso(),
    )
    db.add(log_row)
    db.commit()
    db.refresh(log_row)

    if status != "success":
        raise ValueError(message)

    return {
        "id": log_row.id,
        "status": status,
        "message": message,
        "target_ip": target_ip,
        "target_mac": target_mac,
        "access_switch": locate_data.get("access_switch", {}),
        "interface_name": interface_name,
        "old_vlan": old_vlan,
        "new_vlan": str(new_vlan),
        "operator": log_row.operator,
        "execute_time": log_row.execute_time,
        "warnings": precheck.get("warnings", []),
    }


def execute_vlan_change(db: Session, locate_data: dict, new_vlan: int, operator: str = "unknown") -> dict:
    precheck = _build_precheck_v2(db, locate_data, new_vlan)

    switch_id = int((locate_data.get("access_switch") or {}).get("id") or 0)
    switch = db.query(Device).filter(Device.id == switch_id).first()
    if not switch:
        raise ValueError("鎺ュ叆浜ゆ崲鏈轰笉瀛樺湪")

    interface_name = str(locate_data.get("interface_name") or "").strip()
    if not interface_name:
        raise ValueError("瀹氫綅缁撴灉鏃犳晥锛岀己灏戠粓绔帴鍏ュ彛")
    if "eth-trunk" in interface_name.lower():
        raise ValueError("端口刷新只能执行在终端接入口，不能执行在 Eth-Trunk 口")
    interface_name_config = str(precheck.get("interface_name_config") or "").strip() or _normalize_interface_for_config(interface_name)

    target_ip = str(locate_data.get("target_ip") or "").strip()
    target_mac = _normalize_mac(str(locate_data.get("mac") or "").strip())
    old_vlan = str(locate_data.get("current_vlan") or "").strip()

    executed_steps = []
    verify_steps = []
    status = "failed"
    message = ""
    verify_result = ""
    core_uplink_added = False
    access_uplink_added = False
    vlan_change_success = False
    shutdown_success = False
    undo_shutdown_success = False
    refresh_attempted = False

    try:
        conn = _connect_device(switch, timeout=20)
    except Exception as exc:
        raise ValueError(f"鎺ュ叆浜ゆ崲鏈鸿繛鎺ュけ璐? {exc}") from exc

    try:
        core_switch_id = int((locate_data.get("core_switch") or {}).get("id") or 0)
        core_switch = db.query(Device).filter(Device.id == core_switch_id).first()
        extra_connections = {}
        try:
            for uplink in precheck.get("uplink_plan", []):
                if not uplink.get("requires_add"):
                    continue
                commands = uplink.get("planned_commands") or []
                if not commands:
                    continue

                interface_label = str(uplink.get("interface_name") or "")
                if interface_label.startswith("核心:"):
                    target_device = core_switch
                    connection_key = "core"
                else:
                    target_device = switch
                    connection_key = "access"

                if not target_device:
                    raise ValueError(f"未找到需要补放通的设备: {interface_label}")

                target_conn = conn if connection_key == "access" else extra_connections.get(connection_key)
                if target_conn is None:
                    target_conn = _connect_device(target_device, timeout=20)
                    extra_connections[connection_key] = target_conn

                uplink_out = target_conn.send_config_set(commands, read_timeout=40)
                executed_steps.append(
                    {
                        "device": target_device.name,
                        "stage": "uplink_allow",
                        "interface_name": interface_label,
                        "commands": commands,
                        "output": str(uplink_out or "")[-2000:],
                    }
                )
                if connection_key == "core":
                    core_uplink_added = True
                else:
                    access_uplink_added = True
        finally:
            for key, extra_conn in extra_connections.items():
                if key == "access":
                    continue
                if key == "core" and core_uplink_added:
                    save_out = _huawei_save(extra_conn, timeout=120)
                    executed_steps.append(
                        {
                            "device": core_switch.name if core_switch else "core",
                            "stage": "save",
                            "commands": ["save"],
                            "output": str(save_out or "")[-2000:],
                        }
                    )
                try:
                    extra_conn.disconnect()
                except Exception:
                    pass

        access_change_out = conn.send_config_set(precheck["access_commands"], read_timeout=40)
        vlan_change_success = True
        executed_steps.append(
            {
                "device": switch.name,
                "stage": "access_change",
                "interface_name": interface_name_config,
                "commands": precheck["access_commands"],
                "output": str(access_change_out or "")[-2000:],
            }
        )

        verify_access_cmd = f"display current-configuration interface {interface_name}"
        verify_access_output = _run_show(conn, verify_access_cmd, 20)
        verify_steps.append({"device": switch.name, "command": verify_access_cmd, "output": verify_access_output[-2000:]})
        verify_profile = _parse_port_profile(verify_access_output)
        if str(verify_profile["current_vlan"]) != str(new_vlan):
            raise ValueError("修改后校验失败")

        refresh_attempted = True
        shutdown_commands = [f"interface {interface_name_config}", "shutdown", "quit"]
        shutdown_out = conn.send_config_set(shutdown_commands, read_timeout=20)
        shutdown_success = True
        executed_steps.append(
            {
                "device": switch.name,
                "stage": "port_shutdown",
                "interface_name": interface_name_config,
                "commands": shutdown_commands,
                "output": str(shutdown_out or "")[-2000:],
            }
        )

        undo_shutdown_commands = [f"interface {interface_name_config}", "undo shutdown", "quit"]
        undo_shutdown_out = conn.send_config_set(undo_shutdown_commands, read_timeout=20)
        undo_shutdown_success = True
        executed_steps.append(
            {
                "device": switch.name,
                "stage": "port_undo_shutdown",
                "interface_name": interface_name_config,
                "commands": undo_shutdown_commands,
                "output": str(undo_shutdown_out or "")[-2000:],
            }
        )

        save_out_access = _huawei_save(conn, timeout=120)
        executed_steps.append(
            {
                "device": switch.name,
                "stage": "save",
                "commands": ["save"],
                "output": str(save_out_access or "")[-2000:],
            }
        )

        status = "success"
        message = "VLAN淇敼鎴愬姛锛岀鍙ｅ凡鍒锋柊"
        verify_result = json.dumps(
            {
                "precheck": precheck,
                "verify_steps": verify_steps,
                "core_uplink_added": core_uplink_added,
                "access_uplink_added": access_uplink_added,
                "vlan_change_success": vlan_change_success,
                "shutdown_success": shutdown_success,
                "undo_shutdown_success": undo_shutdown_success,
                "refresh_attempted": refresh_attempted,
                "target_interface": interface_name_config,
                "old_vlan": old_vlan,
                "new_vlan": str(new_vlan),
            },
            ensure_ascii=False,
        )
    except Exception as exc:
        message = str(exc)
        verify_result = json.dumps(
            {
                "precheck": precheck,
                "verify_steps": verify_steps,
                "error": message,
                "core_uplink_added": core_uplink_added,
                "access_uplink_added": access_uplink_added,
                "vlan_change_success": vlan_change_success,
                "shutdown_success": shutdown_success,
                "undo_shutdown_success": undo_shutdown_success,
                "refresh_attempted": refresh_attempted,
                "target_interface": interface_name,
                "old_vlan": old_vlan,
                "new_vlan": str(new_vlan),
            },
            ensure_ascii=False,
        )
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass

    log_row = IpVlanChangeLog(
        operator=str(operator or "unknown"),
        target_ip=target_ip,
        target_mac=target_mac,
        core_switch=str((locate_data.get("core_switch") or {}).get("name") or ""),
        access_switch=str((locate_data.get("access_switch") or {}).get("name") or ""),
        interface_name=interface_name,
        port_type=str(locate_data.get("port_type") or ""),
        port_status=str(locate_data.get("port_status") or ""),
        port_description=str(locate_data.get("port_description") or ""),
        old_vlan=old_vlan,
        new_vlan=str(new_vlan),
        status=status,
        message=message,
        commands=json.dumps(
            {
                "executed_steps": executed_steps,
                "precheck": precheck,
                "old_vlan": old_vlan,
                "new_vlan": str(new_vlan),
                "target_interface": interface_name,
                "refresh_attempted": refresh_attempted,
                "core_uplink_added": core_uplink_added,
                "access_uplink_added": access_uplink_added,
                "shutdown_success": shutdown_success,
                "undo_shutdown_success": undo_shutdown_success,
            },
            ensure_ascii=False,
        ),
        verify_result=verify_result,
        execute_time=_now_iso(),
    )
    db.add(log_row)
    db.commit()
    db.refresh(log_row)

    if status != "success":
        raise ValueError(message)

    return {
        "id": log_row.id,
        "status": status,
        "message": message,
        "target_ip": target_ip,
        "target_mac": target_mac,
        "access_switch": locate_data.get("access_switch", {}),
        "interface_name": interface_name,
        "old_vlan": old_vlan,
        "new_vlan": str(new_vlan),
        "core_uplink_added": core_uplink_added,
        "access_uplink_added": access_uplink_added,
        "vlan_change_success": vlan_change_success,
        "shutdown_success": shutdown_success,
        "undo_shutdown_success": undo_shutdown_success,
        "refresh_attempted": refresh_attempted,
        "operator": log_row.operator,
        "execute_time": log_row.execute_time,
        "warnings": precheck.get("warnings", []),
        "executed_steps": executed_steps,
        "verify_steps": verify_steps,
    }


def list_logs(db: Session, limit: int = 50) -> list[dict]:
    rows = db.query(IpVlanChangeLog).order_by(IpVlanChangeLog.id.desc()).limit(max(1, min(limit, 200))).all()
    return [
        {
            "id": row.id,
            "operator": row.operator,
            "target_ip": row.target_ip,
            "target_mac": row.target_mac,
            "core_switch": row.core_switch,
            "access_switch": row.access_switch,
            "interface_name": row.interface_name,
            "port_type": row.port_type,
            "port_status": row.port_status,
            "port_description": row.port_description,
            "old_vlan": row.old_vlan,
            "new_vlan": row.new_vlan,
            "status": row.status,
            "message": row.message,
            "commands": row.commands,
            "verify_result": row.verify_result,
            "execute_time": row.execute_time,
        }
        for row in rows
    ]


def port_query_locate(
    db: Session,
    query_type: str,
    query_value: str,
    selected_mac: str = "",
    selected_core_uplink: str = "",
) -> dict:
    qt = str(query_type or "ip").strip().lower()
    qv = str(query_value or "").strip()
    if qt not in {"ip", "mac"}:
        raise ValueError("queryType不合法")
    if not qv:
        raise ValueError("queryValue不能为空")

    started = time.time()
    logs: list[dict] = []
    warnings: list[str] = []

    task_id = f"locate_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{int(time.time()*1000)%100000}"

    core = _select_core_switch(db)
    if not core:
        raise LocateTraceError("未找到核心交换机", logs)

    if qt == "ip":
        ip = _validate_ip(qv)
        mac, core_uplink = _port_locate_core_arp(core, ip, logs)
        result = _port_locate_from_mac_and_uplink_stage1(db, mac, core_uplink, logs, target_ip=ip)
    else:
        # MAC fuzzy search on core, may return multiple options to select.
        keyword = _normalize_mac_keyword(qv)
        if not keyword:
            raise ValueError("MAC关键字不能为空")
        match_info = _port_locate_core_mac_candidates(core, keyword, logs)
        if match_info.get("pending_selection"):
            match_info.update(
                {
                    "taskId": task_id,
                    "queryType": "mac",
                    "queryValue": qv,
                    "locateStatus": "PENDING",
                    "warnings": warnings,
                    "durationMs": int((time.time() - started) * 1000),
                    "cached": False,
                }
            )
            return match_info

        mac = match_info["mac"]
        core_uplink = str(selected_core_uplink or match_info.get("core_uplink_interface") or "").strip()
        result = _port_locate_from_mac_and_uplink_stage1(db, mac, core_uplink, logs, target_ip="")

    result.update(
        {
            "taskId": task_id,
            "queryType": qt,
            "queryValue": qv,
            "locateStatus": "SUCCESS",
            "warnings": warnings,
            "durationMs": int((time.time() - started) * 1000),
            "cached": False,
            "logs": logs,
        }
    )
    return result


def port_query_check_trunk(
    db: Session,
    target_vlan: int,
    core_device_id: int,
    core_uplink_port: str,
    access_switch_id: int,
    access_uplink_port: str,
    task_id: str = "",
) -> dict:
    if not isinstance(target_vlan, int) or not (1 <= target_vlan <= 4094):
        raise ValueError("目标VLAN范围需在1~4094之间")
    core_uplink_port = str(core_uplink_port or "").strip()
    access_uplink_port = str(access_uplink_port or "").strip()
    if not core_uplink_port or not access_uplink_port:
        raise ValueError("缺少互联口信息，无法检查放通")

    core = db.query(Device).filter(Device.id == int(core_device_id)).first()
    access = db.query(Device).filter(Device.id == int(access_switch_id)).first()
    if not core:
        raise ValueError("核心交换机不存在")
    if not access:
        raise ValueError("接入交换机不存在")

    started = time.time()
    logs: list[dict] = []

    def fetch_core():
        return _fetch_interface_config(core, core_uplink_port, 20, 60, logs)

    def fetch_access():
        return _fetch_interface_config(access, access_uplink_port, 20, 60, logs)

    with ThreadPoolExecutor(max_workers=2) as executor:
        core_future = executor.submit(fetch_core)
        access_future = executor.submit(fetch_access)
        core_cfg = core_future.result()
        access_cfg = access_future.result()

    core_allowed = _extract_allowed_vlans_text(core_cfg) or "-"
    access_allowed = _extract_allowed_vlans_text(access_cfg) or "-"
    core_has = _interface_vlan_allowed(core_cfg, target_vlan)
    access_has = _interface_vlan_allowed(access_cfg, target_vlan)
    passed = bool(core_has and access_has)

    return {
        "taskId": task_id or f"trunk_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{int(time.time()*1000)%100000}",
        "targetVlan": target_vlan,
        "checkStatus": "SUCCESS" if passed else "FAILED",
        "corePortCheck": {
            "deviceId": core.id,
            "deviceName": core.name,
            "portName": core_uplink_port,
            "portMode": _parse_port_profile(core_cfg)["port_type"],
            "allowed": bool(core_has),
            "allowedVlans": core_allowed,
            "message": "" if core_has else f"核心互联口未放通 VLAN {target_vlan}",
        },
        "accessPortCheck": {
            "deviceId": access.id,
            "deviceName": access.name,
            "portName": access_uplink_port,
            "portMode": _parse_port_profile(access_cfg)["port_type"],
            "allowed": bool(access_has),
            "allowedVlans": access_allowed,
            "message": "" if access_has else f"接入交换机上联口未放通 VLAN {target_vlan}",
        },
        "summary": {"passed": passed, "canModifyDirectly": passed},
        "warnings": [],
        "durationMs": int((time.time() - started) * 1000),
        "logs": logs,
    }


def port_query_change_vlan(
    db: Session,
    task_id: str,
    ip: str,
    mac: str,
    access_switch_id: int,
    access_port: str,
    current_vlan: str,
    target_vlan: int,
    core_device_id: int,
    core_uplink_port: str,
    access_uplink_port: str,
    check_trunk_before_change: bool = True,
    auto_flap_port: bool = True,
    operator: str = "unknown",
) -> dict:
    if not isinstance(target_vlan, int) or not (1 <= target_vlan <= 4094):
        raise ValueError("目标VLAN范围需在1~4094之间")
    if not access_switch_id or not str(access_port or "").strip():
        raise ValueError("缺少接入交换机或端口信息")

    trunk_check_passed = None
    trunk_check = None
    if check_trunk_before_change:
        trunk_check = port_query_check_trunk(
            db,
            target_vlan=target_vlan,
            core_device_id=core_device_id,
            core_uplink_port=core_uplink_port,
            access_switch_id=access_switch_id,
            access_uplink_port=access_uplink_port,
            task_id=task_id,
        )
        trunk_check_passed = bool((trunk_check.get("summary") or {}).get("passed"))
        if not trunk_check_passed:
            raise ValueError("链路 trunk 未放通，禁止修改 VLAN")

    locate_data = {
        "target_ip": str(ip or "").strip(),
        "mac": str(mac or "").strip(),
        "core_switch": {"id": int(core_device_id)},
        "access_switch": {"id": int(access_switch_id)},
        "core_uplink_interface": str(core_uplink_port or "").strip(),
        "access_uplink_interface": str(access_uplink_port or "").strip(),
        "interface_name": str(access_port or "").strip(),
        "current_vlan": str(current_vlan or "").strip(),
        "port_type": "access",
    }

    # Reuse existing execution logic (includes save + flap).
    exec_result = execute_vlan_change(db, locate_data, int(target_vlan), operator=operator)

    command_results = []
    for step in exec_result.get("executed_steps", []) or []:
        if step.get("stage") == "access_change":
            command_results.append({"command": "port default vlan", "success": True, "message": "执行成功"})
        if step.get("stage") == "port_shutdown":
            command_results.append({"command": "shutdown", "success": True, "message": "执行成功"})
        if step.get("stage") == "port_undo_shutdown":
            command_results.append({"command": "undo shutdown", "success": True, "message": "执行成功"})

    return {
        "taskId": task_id or f"change_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{int(time.time()*1000)%100000}",
        "changeStatus": "SUCCESS",
        "accessSwitch": {"deviceId": access_switch_id, "deviceName": (exec_result.get("access_switch") or {}).get("name", "")},
        "accessPort": str(access_port or "").strip(),
        "oldVlan": str(current_vlan or ""),
        "newVlan": int(target_vlan),
        "trunkCheckPassed": True if trunk_check_passed is None else trunk_check_passed,
        "portFlapped": bool(auto_flap_port),
        "commandResults": command_results,
        "durationMs": 0,
        "execute": exec_result,
        "trunkCheck": trunk_check,
    }


def _port_locate_core_arp(core: Device, ip: str, logs: list[dict]) -> tuple[str, str]:
    arp_cache_key = f"arp:{ip}"
    arp_data = _cache_get(arp_cache_key)
    if arp_data is None:
        try:
            conn = _connect_device(core, timeout=20)
        except Exception as exc:
            raise LocateTraceError(f"核心交换机连接失败: {exc}", logs) from exc
        try:
            logs.append({"device": core.name, "command": "ssh-connect", "output": f"SSH登录成功 {core.ip}"})
            arp_cmd = f"display arp | include {ip}"
            arp_output = _run_show(conn, arp_cmd, 20)
            logs.append({"device": core.name, "command": arp_cmd, "output": arp_output[-1500:]})
            if not _parse_arp_mac(arp_output):
                arp_full_cmd = "display arp"
                arp_full_output = _run_show(conn, arp_full_cmd, 30)
                arp_output = _extract_arp_section(arp_full_output, ip)
                logs.append({"device": core.name, "command": arp_full_cmd, "output": (arp_output or "")[-1500:] or "ARP查不到"})
        finally:
            try:
                conn.disconnect()
            except Exception:
                pass
        arp_data = {"mac": _parse_arp_mac(arp_output), "core_uplink": _parse_arp_interface(arp_output) or ""}
        _cache_set(arp_cache_key, arp_data, 45)
    mac = str((arp_data or {}).get("mac") or "").strip()
    core_uplink = str((arp_data or {}).get("core_uplink") or "").strip()
    if not mac:
        raise LocateTraceError("ARP查不到", logs)
    return mac, core_uplink


def _port_locate_core_mac_candidates(core: Device, keyword: str, logs: list[dict]) -> dict:
    try:
        conn = _connect_device(core, timeout=20)
    except Exception as exc:
        raise LocateTraceError(f"核心交换机连接失败: {exc}", logs) from exc
    try:
        logs.append({"device": core.name, "command": "ssh-connect", "output": f"SSH登录成功 {core.ip}"})
        core_cmd = f"display mac-address | include {keyword}"
        core_output = _run_show(conn, core_cmd, 30)
        logs.append({"device": core.name, "command": core_cmd, "output": core_output[-2000:]})
    finally:
        try:
            conn.disconnect()
        except Exception:
            pass

    core_matches = _extract_mac_matches(core_output)
    if not core_matches:
        raise LocateTraceError("核心交换机未匹配到MAC", logs)

    unique_options = []
    seen: set[tuple[str, str]] = set()
    for item in core_matches:
        interface_name = item.get("interface_name") or ""
        key = (item["mac"], interface_name)
        if key in seen:
            continue
        seen.add(key)
        unique_options.append(
            {
                "mac": item["mac"],
                "core_uplink_interface": interface_name,
                "displayText": f"{item['mac']} / {interface_name or '-'}",
            }
        )
    if len(unique_options) > 1:
        return {"pending_selection": True, "match_options": unique_options, "logs": logs}
    return {"pending_selection": False, "mac": unique_options[0]["mac"], "core_uplink_interface": unique_options[0]["core_uplink_interface"]}


def _port_locate_from_mac_and_uplink_stage1(db: Session, mac: str, core_uplink: str, logs: list[dict], target_ip: str = "") -> dict:
    core = _select_core_switch(db)
    if not core:
        raise LocateTraceError("未找到核心交换机", logs)

    core_uplink = str(core_uplink or "").strip()
    topology_row = _find_topology_row(core_uplink)
    topology_switch = _find_access_switch_by_row(db, topology_row)
    hinted_switch = topology_switch

    access_uplink = _choose_access_uplink_port(topology_row, fallback_port=core_uplink)

    candidate_switches = _ordered_access_switches(db, target_ip, hinted_switch)
    matches = []
    for switch in [item for item in candidate_switches if item]:
        try:
            conn = _connect_device(switch, timeout=30)
        except Exception as exc:
            logs.append({"device": switch.name, "command": "connect", "output": f"SSH connect failed: {exc}"})
            continue
        try:
            logs.append({"device": switch.name, "command": "ssh-connect", "output": f"SSH登录成功 {switch.ip}"})
            for mac_cmd in _build_mac_lookup_commands(mac):
                mac_output = _run_show(conn, mac_cmd, 12)
                logs.append({"device": switch.name, "command": mac_cmd, "output": mac_output[-1500:]})
                interface_name = _parse_port_from_mac(mac_output)
                if interface_name:
                    matches.append({"switch": switch, "interface_name": interface_name})
                    break
        finally:
            try:
                conn.disconnect()
            except Exception:
                pass

    if not matches:
        raise LocateTraceError("MAC查不到", logs)
    if len(matches) > 1:
        names = ", ".join(f"{item['switch'].name}:{item['interface_name']}" for item in matches[:5])
        raise LocateTraceError(f"找到多个端口，无法唯一定位: {names}", logs)

    access_switch: Device = matches[0]["switch"]
    access_port = matches[0]["interface_name"]

    try:
        cfg_output = _fetch_interface_config(access_switch, access_port, 20, 60, logs)
    except Exception as exc:
        raise LocateTraceError(f"读取接入口配置失败: {exc}", logs) from exc

    profile = _parse_port_profile(cfg_output)
    forbidden, forbidden_reason = _is_forbidden_port(access_port, profile)
    if forbidden:
        raise LocateTraceError(forbidden_reason, logs)
    if profile["port_type"] != "access":
        raise LocateTraceError("当前端口不是 access 口", logs)

    return {
        "ip": str(target_ip or ""),
        "mac": _normalize_mac(mac),
        "accessSwitch": {"deviceId": access_switch.id, "deviceName": access_switch.name, "managementIp": access_switch.ip},
        "accessPort": access_port,
        "currentVlan": int(profile["current_vlan"]),
        "coreDevice": {"deviceId": core.id, "deviceName": core.name, "managementIp": core.ip},
        "coreUplinkPort": core_uplink,
        "accessUplinkPort": access_uplink,
        "pathType": "core_to_access",
    }
