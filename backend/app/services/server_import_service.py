from __future__ import annotations

import csv
import io
import ipaddress
import uuid
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.server import ServerAsset
from app.services.server_service import _normalize_server_payload

PREVIEW_TTL_MINUTES = 30
_PREVIEW_STORE: dict[str, dict] = {}

HEADER_MAP = {
    "name": "name",
    "名称": "name",
    "服务器名称": "name",
    "ip": "ip",
    "ip地址": "ip",
    "主机地址": "ip",
    "hostname": "hostname",
    "host": "hostname",
    "主机名": "hostname",
    "server_type": "server_type",
    "type": "server_type",
    "服务器类型": "server_type",
    "access_method": "access_method",
    "access": "access_method",
    "接入方式": "access_method",
    "username": "username",
    "user": "username",
    "用户名": "username",
    "password": "password",
    "pwd": "password",
    "密码": "password",
    "port": "port",
    "端口": "port",
    "group_name": "group_name",
    "group": "group_name",
    "所属分组": "group_name",
    "enable": "enable",
    "enabled": "enable",
    "启用": "enable",
}

REQUIRED_FIELDS = ["name", "ip", "server_type", "username", "password"]


def _cleanup_expired_previews() -> None:
    now = datetime.now()
    expired = [
        key
        for key, item in _PREVIEW_STORE.items()
        if now - item.get("created_at", now) > timedelta(minutes=PREVIEW_TTL_MINUTES)
    ]
    for key in expired:
        _PREVIEW_STORE.pop(key, None)


def _normalize_header(header: str) -> str:
    return (
        (header or "")
        .replace("\ufeff", "")
        .replace("\u3000", "")
        .strip()
        .lower()
        .replace(" ", "")
    )


def _map_header(header: str) -> str | None:
    normalized = _normalize_header(header)
    if not normalized:
        return None
    if normalized in HEADER_MAP:
        return HEADER_MAP[normalized]
    fuzzy_rules = [
        (["name", "名称", "服务器名称"], "name"),
        (["ip", "ip地址", "主机地址"], "ip"),
        (["hostname", "host", "主机名"], "hostname"),
        (["server_type", "type", "服务器类型"], "server_type"),
        (["access_method", "access", "接入方式"], "access_method"),
        (["username", "user", "用户名"], "username"),
        (["password", "pwd", "密码"], "password"),
        (["port", "端口"], "port"),
        (["group_name", "group", "所属分组"], "group_name"),
        (["enable", "enabled", "启用"], "enable"),
    ]
    for keys, target in fuzzy_rules:
        if any(key in normalized for key in keys):
            return target
    return None


def _normalize_row(raw_row: dict) -> dict:
    normalized = {
        "name": "",
        "ip": "",
        "hostname": "",
        "server_type": "linux",
        "access_method": "",
        "username": "",
        "password": "",
        "port": "",
        "group_name": "default",
        "enable": 1,
    }
    for key, value in raw_row.items():
        mapped = _map_header(str(key))
        if not mapped:
            continue
        normalized[mapped] = "" if value is None else str(value).strip()
    if normalized["enable"] == "":
        normalized["enable"] = 1
    return normalized


def _parse_enable(value) -> int:
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
        return 1
    if text in {"0", "false", "no", "n", "off"}:
        return 0
    raise ValueError("启用字段非法，应为 0/1/true/false")


def _validate_row(row: dict) -> tuple[bool, list[str], dict]:
    errors: list[str] = []
    payload = row.copy()

    for field in REQUIRED_FIELDS:
        if not str(payload.get(field, "")).strip():
            errors.append(f"必填字段缺失: {field}")

    ip_value = str(payload.get("ip", "")).strip()
    if ip_value:
        try:
            ipaddress.ip_address(ip_value)
        except Exception:
            errors.append("IP地址格式非法")

    server_type = str(payload.get("server_type", "")).strip().lower()
    if server_type and server_type not in {"linux", "windows"}:
        errors.append("服务器类型仅支持 Linux / Windows")

    port_val = payload.get("port", "")
    if str(port_val).strip():
        try:
            port_int = int(str(port_val))
            if port_int < 1 or port_int > 65535:
                raise ValueError
            payload["port"] = port_int
        except Exception:
            errors.append("端口非法，应为 1-65535")

    try:
        payload["enable"] = _parse_enable(payload.get("enable", 1))
    except Exception as exc:
        errors.append(str(exc))

    if not payload.get("group_name"):
        payload["group_name"] = "default"

    normalized_payload = _normalize_server_payload(payload)
    return (len(errors) == 0), errors, normalized_payload


def _read_csv_rows(content: bytes) -> list[dict]:
    for enc in ["utf-8-sig", "utf-8", "gbk", "cp936", "latin-1"]:
        try:
            text = content.decode(enc)
            reader = csv.DictReader(io.StringIO(text))
            return list(reader)
        except Exception:
            continue
    raise ValueError("CSV 文件编码无法识别")


def _read_excel_rows(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = []
    for values in rows[1:]:
        row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
        if any(str(v).strip() for v in row.values() if v is not None):
            data_rows.append(row)
    return data_rows


def preview_import(db: Session, filename: str, content: bytes) -> dict:
    _cleanup_expired_previews()

    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        raw_rows = _read_csv_rows(content)
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raw_rows = _read_excel_rows(content)
    else:
        raise ValueError("仅支持 CSV 或 Excel(.xlsx) 文件")

    existing_ips = {ip for (ip,) in db.query(ServerAsset.ip).all()}
    seen_ips: set[str] = set()

    preview_rows = []
    importable_payloads = []
    summary = {"total": 0, "importable": 0, "exists": 0, "duplicate": 0, "error": 0}

    for idx, raw in enumerate(raw_rows, start=1):
        normalized = _normalize_row(raw)
        valid, errors, payload = _validate_row(normalized)

        status = "importable"
        reason: list[str] = []
        ip = str(payload.get("ip", "")).strip()

        if not valid:
            status = "error"
            reason.extend(errors)
        elif ip in seen_ips:
            status = "duplicate"
            reason.append("文件内 IP 地址重复")
        elif ip in existing_ips:
            status = "exists"
            reason.append("数据库已存在该 IP，已跳过")

        if ip:
            seen_ips.add(ip)

        summary["total"] += 1
        if status == "importable":
            summary["importable"] += 1
            importable_payloads.append(payload)
        elif status == "exists":
            summary["exists"] += 1
        elif status == "duplicate":
            summary["duplicate"] += 1
        else:
            summary["error"] += 1

        preview_rows.append(
            {
                "row_no": idx,
                "status": status,
                "reason": "；".join(reason) if reason else "可导入",
                "server": payload,
            }
        )

    preview_id = uuid.uuid4().hex
    _PREVIEW_STORE[preview_id] = {
        "created_at": datetime.now(),
        "rows": preview_rows,
        "importable_payloads": importable_payloads,
        "summary": summary,
    }
    return {"preview_id": preview_id, "summary": summary, "rows": preview_rows}


def confirm_import(db: Session, preview_id: str) -> dict:
    _cleanup_expired_previews()
    item = _PREVIEW_STORE.get(preview_id)
    if not item:
        raise ValueError("预检查记录不存在或已过期，请重新上传")

    payloads = item.get("importable_payloads", [])
    imported = 0
    existing_ips = {ip for (ip,) in db.query(ServerAsset.ip).all()}

    for payload in payloads:
        ip = payload.get("ip")
        if ip in existing_ips:
            continue
        db.add(ServerAsset(**payload))
        existing_ips.add(ip)
        imported += 1

    db.commit()
    skipped = item.get("summary", {}).get("total", 0) - imported
    _PREVIEW_STORE.pop(preview_id, None)
    return {"imported": imported, "skipped": skipped}
